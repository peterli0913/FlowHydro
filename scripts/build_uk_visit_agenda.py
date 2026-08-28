"""
Build the UK (Sandwich) visit agenda as a new sheet that reuses the exact table
style of the "Agenda" sheet in the SW workbook.

The source workbook is copied first and never modified, and the new sheet is
created inside that copy so the banner fill (a theme colour) resolves to the
same colour as in the original.

Plan encoded here: site walkdown 0.5 d, 3D model review 2.5 d, operating
manual / control philosophy / sequence control 2 d, HAZOP 3 d -> Day 1 to
Day 8.
"""

import shutil
from copy import copy

import openpyxl

SRC = "【SW】Agenda of the visit 0130 V4.xlsx"
DST = "Agenda of the visit to Sandwich (3D & HAZOP).xlsx"
NEW_SHEET = "UK Visit Agenda"

PERSONS = "Keith / 李涛 / 范双双 / 胡超群 / 孟德智"

HEADERS = ["No.", "Date", "Time", "Itinerary", "Person in Charge",
           "Detailed Topic", "Location", "Objective  目标"]

AM, PM, LUNCH = "09:00-12:00", "14:00-17:00", "12:00-13:00"

# (banner, [(time, itinerary, detail, objective) ...])  — one entry per day
DAYS = [
    ("Site Walkdown & 3D Model Review", [
        (AM,
         "1. 现场勘察\nSite Walkdown\n（0.5 天 / 0.5 day）",
         "1, Skid boundary line, stairwell clearance and the faces actually "
         "available for access.\n"
         "2, Emergency shower and eyewash — actual position and dimensions.\n"
         "3, HTF wall-penetration position and the equipment-room interface.\n"
         "4, Goods lift and delivery route dimensions.\n"
         "现场核对：撬块边界线与楼梯井净空、可用操作面；紧急淋浴与洗眼器实际位置与尺寸；"
         "HTF 穿墙位置与设备间接口；货梯与进场路径尺寸。",
         "Measured constraints recorded and jointly confirmed on site, forming "
         "the fixed basis for every layout decision made this week.\n"
         "现场实测约束当场记录并双方确认，作为本周全部布置决策的固定基准。"),
        (LUNCH, "Lunch", None, None),
        (PM,
         "2.1 设备顺序调整\nEquipment Sequence Re-arrangement\n（0.5 天 / 0.5 day）",
         "1, Whether the U-shape re-sequencing releases the estimated "
         "500–600 mm.\n"
         "2, Positions of preheater, condenser, separator and product tank.\n"
         "3, Pipe routing length and crossings after re-sequencing.\n"
         "核算 U 形重排实际可释放的宽度；预热器 / 冷凝器 / 分离器 / 产品罐位置；"
         "重排后管路长度与交叉情况。",
         "Equipment sequence fixed and the released width agreed as a figure; "
         "model updated the same day.\n"
         "设备顺序定案，释放宽度以数值形式双方认可，当日更新模型。"),
    ]),

    ("3D Model Review", [
        (AM,
         "2.2 维修通道的变更\nMaintenance Access Revision\n（0.5 天 / 0.5 day）",
         "1, Central walkway versus removable module — one route to be chosen.\n"
         "2, Manual handling of flanges at approx. 100 kg per item inside a "
         "confined space.\n"
         "3, Flange-to-wall distance, reach-in depth and gasket replacement "
         "space.\n"
         "中间通道与可拆卸模块二选一；单件约 100 kg 法兰在受限空间内的人工搬运；"
         "法兰离墙距离、探身深度与垫片更换所需空间。",
         "Maintenance access route decided, together with its stated impact on "
         "site disassembly workload and SAT scope.\n"
         "维护可达性路线定案，并明确其对现场拆装工程量与 SAT 范围的影响。"),
        (LUNCH, "Lunch", None, None),
        (PM,
         "2.3 控制柜布置（分体）调整方案\nControl Panel Layout / Split Scheme\n"
         "（1 天 / 1 day — 第 1 段 Part 1）",
         "1, Nature of the clash with the emergency shower — physical "
         "interference or water spray.\n"
         "2, Whether the eyewash can be relocated to just outside the door.\n"
         "3, IS and non-IS panels must be split as separate enclosures.\n"
         "控制柜与紧急淋浴冲突的性质（物理干涉或淋浴溅水）；洗眼器能否移至门外；"
         "本安柜与非本安柜须分别拆分。",
         "Clash nature confirmed and the eyewash relocation answered, so the "
         "panel position ceases to be an open loop.\n"
         "冲突性质确认、洗眼器能否外移给出明确答复，控制柜位置不再悬置。"),
    ]),

    ("3D Model Review", [
        (AM,
         "2.3 控制柜布置（分体）调整方案\nControl Panel Layout / Split Scheme\n"
         "（第 2 段 Part 2）",
         "1, Junction box split and cable count alignment — 16 against 21.\n"
         "2, Hinged panel so tested cabling need not be disconnected for "
         "shipping.\n"
         "3, Panel sizes computed against the instrument allocation agreed in "
         "this session.\n"
         "接线箱拆分方案与电缆数对齐（16 根与 21 根的差异）；柜体铰接方案实现免拆已测电缆发运；"
         "按本段确定的仪表归属核算柜体尺寸。",
         "Panel quantity, size and mounting position fixed; cable count "
         "reconciled to a single agreed figure.\n"
         "柜体数量、尺寸与安装位置定案，电缆数对齐为唯一双方认可的数值。"),
        (LUNCH, "Lunch", None, None),
        (PM,
         "2.4 四个分体撬块分开方案\nFour-Module Split Scheme\n（0.5 天 / 0.5 day）",
         "1, Split boundaries and the I/O belonging to each module.\n"
         "2, Lifting lugs, skates handling and the re-assembly sequence on "
         "site.\n"
         "3, Overall height for the goods lift — top section removal, HTF "
         "pipework shipped loose.\n"
         "分体边界与各模块 I/O 归属；吊耳、skates 搬运方式与现场重组顺序；"
         "整撬高度适配货梯（去掉顶部一段、HTF 管道散件发运）。",
         "Split boundaries and per-module I/O frozen; the height solution for "
         "the goods lift confirmed feasible.\n"
         "分体边界与各模块 I/O 冻结，进货梯的降高方案确认可行。"),
    ]),

    ("Operating Manual", [
        (AM,
         "3.1 操作手册讨论\nOperating Manual Review\n"
         "（1 天 / 1 day — 第 1 段 Part 1）",
         "1, Walk through the manual section by section against the layout "
         "fixed earlier this week.\n"
         "2, Catalyst charging and discharging, including the water-wet "
         "discharge sequence.\n"
         "3, Cleaning operations — valve line-up and hose connections.\n"
         "结合本周已定案的布置逐节走查操作手册；催化剂装料与卸料（含水润催化剂卸料顺序）；"
         "清洗操作的阀门开关与软管连接。",
         "Comments captured and dispositioned within the session; the manual "
         "revision scope is agreed.\n"
         "意见当场记录并逐条给出处置，操作手册的修订范围达成一致。"),
        (LUNCH, "Lunch", None, None),
        (PM,
         "3.1 操作手册讨论\nOperating Manual Review\n（第 2 段 Part 2）",
         "1, Start-up, normal operation, shutdown and emergency response.\n"
         "2, Operator access to valves and sampling points, checked against "
         "the walkdown findings.\n"
         "3, Remaining comments closed out one by one.\n"
         "开车、正常操作、停车与应急处置；结合现场勘察结果确认阀门与取样点的操作可达性；"
         "剩余意见逐条关闭。",
         "Operating manual signed off, or every remaining comment carries an "
         "agreed resolution and owner.\n"
         "操作手册签认，或每条未决意见均已有双方认可的处置方案与责任人。"),
    ]),

    ("Control Philosophy & Sequence Control", [
        (AM,
         "3.2 控制说明讨论\nControl Philosophy Review\n（0.5 天 / 0.5 day）",
         "1, Control loops, interlocks and fail positions.\n"
         "2, Division between hardwired interlocks and remote I/O.\n"
         "3, Instrument signal types — Modbus or 4–20 mA, IS or Ex d.\n"
         "控制回路、联锁与故障位；硬联锁与远程 I/O 的划分；"
         "仪表信号形式（Modbus 或 4–20 mA、本安或隔爆）。",
         "Control philosophy agreed and the instrument selection items closed, "
         "so the I/O schedule can be issued.\n"
         "控制说明达成一致，仪表选型条目关闭，I/O 表可据此发布。"),
        (LUNCH, "Lunch", None, None),
        (PM,
         "3.3 顺控方案讨论\nSequence Control Review\n（0.5 天 / 0.5 day）",
         "1, Sequence steps, permissives and hold points.\n"
         "2, Cleaning sequence including inerting, and the high-level "
         "interlock bypass in cleaning mode.\n"
         "3, Operator interface and alarm handling.\n"
         "顺控步骤、允许条件与保持点；清洗顺控（含惰化）及清洗模式下高液位联锁的旁通；"
         "操作界面与报警处理。",
         "Sequence scheme agreed as an input to HAZOP, with no sequence "
         "question left open going into the study.\n"
         "顺控方案达成一致并作为 HAZOP 的输入，进入分析前不留未定的顺控问题。"),
    ]),

    ("HAZOP Study", [
        (AM,
         "4. HAZOP 分析讨论\nHAZOP Study\n（3 天 / 3 days — 第 1 天 Day 1）",
         "1, Confirm scope, node split, and the P&ID and layout revision to be "
         "studied.\n"
         "2, Nodes: feed, buffer tanks BT01 / BT04 and metering pumps.\n"
         "确认分析范围、节点划分，以及所用 P&ID 与布置的版本；"
         "节点：进料、缓冲罐 BT01 / BT04 与计量泵。",
         "Study basis locked to the layout fixed this week; the day's nodes "
         "completed with actions and owners recorded in the meeting.\n"
         "分析基准锁定为本周定案的布置；当日节点全部完成，行动项与责任人会上直接记录。"),
        (LUNCH, "Lunch", None, None),
        (PM,
         "4. HAZOP 分析讨论\nHAZOP Study\n（第 1 天 续 Day 1 cont.）",
         "Nodes: preheater and reactors CR01–CR03, covering cooling failure, "
         "hot spot and hydrogen supply.\n"
         "节点：预热器与反应器 CR01–CR03，涵盖冷却失效、热点与氢气供给工况。",
         "Nodes closed out with recommendations agreed on the spot rather than "
         "deferred.\n"
         "节点当场关闭，建议措施当场达成一致，不作延后处理。"),
    ]),

    ("HAZOP Study", [
        (AM,
         "4. HAZOP 分析讨论\nHAZOP Study\n（第 2 天 Day 2）",
         "Nodes: condenser, gas-liquid separator SE01, relief and vent "
         "system.\n"
         "节点：冷凝器、气液分离器 SE01、安全泄放与放空系统。",
         "Nodes closed out with recommendations and owners agreed in the "
         "session.\n"
         "节点当场关闭，建议措施与责任人会上确定。"),
        (LUNCH, "Lunch", None, None),
        (PM,
         "4. HAZOP 分析讨论\nHAZOP Study\n（第 2 天 续 Day 2 cont.）",
         "Nodes: product buffer tank, polishing filter and product discharge; "
         "valve accessibility and sampling reviewed against the walkdown.\n"
         "节点：产品缓冲罐、精滤器与产品出料；结合现场勘察复核阀门可达性与取样方式。",
         "Any layout or valve change arising is resolved against the physical "
         "space in the same session.\n"
         "由此引出的布置或阀门变更，在同一时段内对照实体空间解决。"),
    ]),

    ("HAZOP Study", [
        (AM,
         "4. HAZOP 分析讨论\nHAZOP Study\n（第 3 天 Day 3）",
         "Nodes: CIP and cleaning circuit, inerting, utilities, control and "
         "interlock functions.\n"
         "节点：CIP 与清洗回路、惰化、公用工程、控制与联锁功能。",
         "All nodes completed; no node left for a follow-up study.\n"
         "全部节点完成，不遗留需另行安排的节点。"),
        (LUNCH, "Lunch", None, None),
        (PM,
         "Close-out\n结论汇总与签认\n（第 3 天 续 Day 3 cont.）",
         "1, Review the full action list, assign owners and dates, and sign "
         "off both sides.\n"
         "2, Confirm which documents require revision and the issue dates.\n"
         "3, Record the decisions taken this week against the tracker items "
         "they close.\n"
         "逐条复核行动项清单、确定责任人与时间并双方签认；确认需要升版的文件及发布日期；"
         "将本周形成的决策对应到其所关闭的跟踪表条目。",
         "Signed action list and revision list issued before departure, with "
         "every item of this agenda either closed or carrying an agreed "
         "resolution.\n"
         "离场前形成经双方签认的行动项与文件升版清单，本议程各项均已关闭或已有双方认可的处置方案。"),
    ]),
]

TOTAL_NOTE = ("Total 合计：8 days / 8 天　—　Site walkdown 0.5 + 3D model 2.5 "
              "+ Operating manual & control 2 + HAZOP 3　（现场勘察 0.5 + "
              "3D 模型 2.5 + 操作手册与控制说明 2 + HAZOP 3）")


def style_from(src_cell, dst_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format


def main():
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    src = wb["Agenda"]

    # style templates taken from the original sheet
    T_TITLE = src["A1"]
    T_HDR = src["A2"]
    T_BANNER = src["A5"]
    T_NO = src["A6"]              # No. cell — integer number format
    T_CENTER = src["C6"]          # centred data cell
    T_LEFT = src["D6"]            # left-aligned data cell

    if NEW_SHEET in wb.sheetnames:
        del wb[NEW_SHEET]
    ws = wb.create_sheet(NEW_SHEET, 0)

    last_col = len(HEADERS)                      # 8 -> column H
    last_letter = openpyxl.utils.get_column_letter(last_col)

    for letter, dim in src.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[letter].width = dim.width
    ws.column_dimensions[last_letter].width = 42

    # ---- title -------------------------------------------------------
    ws.cell(row=1, column=1,
            value="【SW】Agenda of the visit to Sandwich — 3D Layout & HAZOP")
    for col in range(1, last_col + 1):
        style_from(T_TITLE, ws.cell(row=1, column=col))
    ws.row_dimensions[1].height = src.row_dimensions[1].height

    # ---- header ------------------------------------------------------
    for col, text in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=col, value=text)
        style_from(T_HDR, c)

    # ---- body --------------------------------------------------------
    row = 3
    no = 1
    for day_index, (banner, items) in enumerate(DAYS, start=1):
        ws.cell(row=row, column=1, value=banner)
        for col in range(1, last_col + 1):
            style_from(T_BANNER, ws.cell(row=row, column=col))
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=last_col)
        row += 1

        day_first = row
        for time, itinerary, detail, objective in items:
            ws.cell(row=row, column=1, value=no)
            ws.cell(row=row, column=2, value=None)
            ws.cell(row=row, column=3, value=time)
            ws.cell(row=row, column=4, value=itinerary)
            ws.cell(row=row, column=5, value=None if detail is None else PERSONS)
            ws.cell(row=row, column=6, value=detail)
            ws.cell(row=row, column=7, value=None)
            ws.cell(row=row, column=8, value=objective)
            style_from(T_NO, ws.cell(row=row, column=1))
            for col in (2, 3):
                style_from(T_CENTER, ws.cell(row=row, column=col))
            for col in (4, 5, 6, 7, 8):
                style_from(T_LEFT, ws.cell(row=row, column=col))
            no += 1
            row += 1

        ws.cell(row=day_first, column=2, value=f"Day {day_index}")
        ws.merge_cells(start_row=day_first, start_column=2,
                       end_row=row - 1, end_column=2)

    # ---- total -------------------------------------------------------
    ws.cell(row=row, column=1, value=TOTAL_NOTE)
    for col in range(1, last_col + 1):
        style_from(T_BANNER, ws.cell(row=row, column=col))
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=last_col)

    ws.freeze_panes = "A3"
    wb.save(DST)
    print(f"Saved: {DST}")
    print(f"sheet '{NEW_SHEET}': rows={row}, cols={last_col}")


if __name__ == "__main__":
    main()
