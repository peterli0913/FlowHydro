"""Build a self-contained HTML visualization of the Sandwich
Continuous Hydrogenation project timeline.

Inputs (kept in repo root):
  - 0509-Sandwich Continuous Hydrogen Timeline.xlsx
  - 05-09 内部会议. 控制策略与防爆分区讨论.docx
  - 2026-05-09 15.47 录音.docx

Output:
  - timeline_view.html (committed alongside the source files so it can be
    opened directly in a browser)

The script extracts the colored Gantt cells from the "Timeline " sheet and
renders both the original (blue) and the revised / current (orange) plans
on the same chart, overlaid with "需要更新的时间节点" (purple) and "CE 认证
公司介入节点" (green). Yellow cells in the xlsx are intentionally not
drawn. The script also encodes the decisions / actions / risks from the
2026-05-09 meetings (rendered above the chart, not as fake bars).
"""

from __future__ import annotations

import html
import json
import os
from dataclasses import dataclass
from datetime import date, timedelta
from typing import Dict, List, Optional, Tuple

import openpyxl

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
XLSX_PATH = os.path.join(REPO_ROOT, "0509-Sandwich Continuous Hydrogen Timeline.xlsx")
HTML_PATH = os.path.join(REPO_ROOT, "timeline_view.html")

# Calendar anchor: column E (the first month/week column) corresponds to
# the first week of August 2025. Each subsequent column is one project
# "week" inside the same month.
CAL_ANCHOR_YEAR = 2025
CAL_ANCHOR_MONTH = 8
TODAY = date(2026, 5, 9)

# Color buckets ----------------------------------------------------------------
COLOR_ORIGINAL = "FF00B0F0"      # 浅蓝 = 原计划
COLOR_REVISED = "FFFFC000"       # 橙黄 = 本次更新后的计划
COLOR_UPDATE = "FF7030A0"        # 紫 = 需要更新的时间节点
COLOR_CERT_GREEN = "FF92D050"    # 绿 = CE 认证公司介入节点
COLOR_CERT_DARKGREEN = "FF00B050"
COLOR_HAZOP_IGNORE = "FFFFFF00"  # 黄 = 不绘制（按 5/11 反馈去掉）

PALETTE = {
    "original": "#7DCBED",
    "revised":  "#F0A93B",
    "update":   "#7C3FAA",
    "cert":     "#5DB851",
}


@dataclass
class Cell:
    col: int                # absolute column index in xlsx
    bucket: str             # original | revised | update | cert | note
    note: Optional[str]     # any text in the cell


@dataclass
class Item:
    row: int
    section: str            # Design / Skid / On-Site Engineering / On-Site Install
    name_en: str
    owner: str
    cells: List[Cell]


def classify_color(rgb: Optional[str]) -> Optional[str]:
    if not rgb:
        return None
    rgb = rgb.upper()
    if rgb == COLOR_ORIGINAL:
        return "original"
    if rgb == COLOR_REVISED:
        return "revised"
    if rgb == COLOR_UPDATE:
        return "update"
    if rgb in (COLOR_CERT_GREEN, COLOR_CERT_DARKGREEN):
        return "cert"
    # Yellow (HAZOP) and any other colors -> intentionally ignored
    return None


def col_to_month_week(col_idx: int) -> Tuple[int, int, int]:
    """Return (year, month, week_in_month 1..4) for a Timeline-sheet column.

    The Timeline sheet starts the calendar at column E (idx 5) which we
    anchor to Aug 2025 wk 1.
    """
    offset = col_idx - 5      # 0-based week offset from Aug 2025 wk1
    if offset < 0:
        return (0, 0, 0)
    month_offset = offset // 4
    week = offset % 4 + 1
    m = CAL_ANCHOR_MONTH + month_offset
    y = CAL_ANCHOR_YEAR + (m - 1) // 12
    m = (m - 1) % 12 + 1
    return (y, m, week)


def col_to_approx_date(col_idx: int) -> date:
    """Roughly map a Timeline column to a real date (Monday of that
    week-of-month). Used only for sorting and the 'today' line."""
    y, m, w = col_to_month_week(col_idx)
    if y == 0:
        return date(CAL_ANCHOR_YEAR, CAL_ANCHOR_MONTH, 1)
    day = min(28, 1 + (w - 1) * 7)
    return date(y, m, day)


def date_to_col(target: date) -> float:
    """Inverse: return a fractional column index for a real date so we can
    draw a 'today' marker on the chart."""
    months_from_anchor = (target.year - CAL_ANCHOR_YEAR) * 12 + (target.month - CAL_ANCHOR_MONTH)
    week_in_month = min(3.99, max(0, (target.day - 1) / 7.0))
    return 5 + months_from_anchor * 4 + week_in_month


# ---- read xlsx ---------------------------------------------------------------

def load_items() -> Tuple[List[Item], int, int]:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Timeline "]
    max_col = ws.max_column
    items: List[Item] = []

    # Row 4..54 contains work items; section header tags appear in col B
    current_section = ""
    for r in range(4, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if b:
            current_section = str(b).strip()
        name = ws.cell(row=r, column=3).value
        owner = ws.cell(row=r, column=4).value
        if not name:
            continue
        cells: List[Cell] = []
        for c in range(5, max_col + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            fg = None
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                fg = cell.fill.fgColor.rgb
            bucket = classify_color(fg)
            note = None
            if isinstance(v, str):
                note = v.strip()
            elif v is not None:
                note = str(v)
            if bucket or note:
                if bucket is None:
                    # keep note-only cells too (rare)
                    bucket = "note"
                cells.append(Cell(col=c, bucket=bucket, note=note))
        items.append(Item(row=r, section=current_section, name_en=str(name).strip(),
                          owner=str(owner).strip() if owner else "", cells=cells))
    return items, 5, max_col


# ---- chinese localisation ----------------------------------------------------

ITEM_LABEL_ZH = {
    "HAZOP （support resource in grey）": "HAZOP（灰色为辅助资源）",
    "Control Description": "控制说明书 (Control Description)",
    "PFD and Mass Balance Data Sheet": "PFD 与物料平衡数据表",
    "PI&D": "PI&D 管道仪表流程图",
    "Process Equipment List": "工艺设备一览表",
    "Equipment datasheets": "设备数据单 (Equipment datasheets)",
    "Battery Limit Condition Sheet (Tie-in point sheet)": "界区条件表 / Tie-in 点表",
    "Instrumentation Condition Table": "仪表条件表",
    "Piping Material Class Specification": "管道材料等级规定",
    "Process Piping List": "工艺管线表",
    "Special Piping Accessories Data Sheet": "特殊管件数据表",
    "Main Safety Relief Facilities Data Sheet": "主要安全泄放设施数据表",
    "Instrument datasheets": "仪表数据单",
    "Interlock logic": "联锁逻辑",
    "Pump specification sheet": "泵规格书",
    "Operating instructions": "操作手册",
    "Preliminary equipment layout": "初步设备布置图",
    "structural design of static equipment（pink）": "静设备结构设计（pink）",
    "Reactor column specification sheet": "反应柱规格书",
    "Procurement information confirmation(pump\\insturment\\equipment schedule)":
        "采购信息确认（泵 / 仪表 / 设备清单）— 含 Keith 选型确认",
    "FDS(CS in FDS)/SDS/HDS": "FDS / SDS / HDS（含控制策略 CS）",
    "URS": "URS 用户需求文件",
    "Instrument Schedule  (including instrument ranges and set values)":
        "仪表清单（含量程与设定值）",
    "I/O schedule (including I/O type, ranges and alarm settings)":
        "I/O 清单（含 I/O 类型、量程、报警设定）",
    "trip and interlock schedules for PLC interlocks": "PLC 联锁 Trip & Interlock 表",
    "PCS interlocks and hardwired interlocks": "PCS 联锁 + 硬连锁",
    "SIL assessments for safety/environmental interlocks and LOPA assessments for potential SIL rated SIFs":
        "SIL / LOPA 评估（安全/环境联锁、SIL 等级 SIF）",
    "Alarm list": "报警清单",
    "Hardware drawing of PLC": "PLC 硬件图",
    "IOA of PLC": "PLC IO 分配 (IOA)",
    "Cable schedule": "电缆清册 (Cable schedule)",
    "3D modelling (procurement feedback; UK confirmation of equipment layout)":
        "3D 建模（采购反馈 + UK 确认布置）",
    "BOM": "BOM 物料清单",
    "Fabrication drawings": "制造装配图 (Fabrication drawings)",
    "Engineering procurement (valves & instruments, static equipment, pumps, piping and fittings; within the skid)":
        "工程采购（阀门仪表 / 静设备 / 泵 / 管件 — 撬内）",
    "Control system procurement and configuration": "控制系统采购与组态（含柜体 ATEX）",
    "Equipment fabrication, assembly and debugging": "设备加工、组装与调试",
    "Integrated factory acceptance (FAT)": "整撬工厂验收 FAT",
    "CE certification (individual equipment + whole-skid certification per Notified Body requirements); full documentation completed":
        "CE 认证（单台 + 整撬，按 NB 要求；文件齐套）",
    "Export formalities, transport inspection/certification and packing":
        "出口手续 / 运输检验 / 包装",
    "Export shipment (Europe)": "出口运输至欧洲",
    "PID": "PID（现场端）",
    "Engineering design documents": "工程设计文件（现场端）",
    "HAZOP": "HAZOP（现场端，UK 主导）",
    "Detailed design": "详细设计（现场端）",
    "Procurement": "采购（现场端 — 撬外件）",
    "Off-skid/peripheral installation": "撬外 / 外围安装",
    "Skid installation": "撬块安装",
    "Mechanical, electrical, instrumentation & piping (MEIP) installation":
        "MEIP 机/电/仪/管安装",
    "Qualification & acceptance testing (DQ / IQ / OQ)":
        "确认与验收测试 (DQ / IQ / OQ)",
    "Commissioning and start-up": "调试与开车",
}

SECTION_LABEL_ZH = {
    "Design": "设计阶段（撬块文件输出）",
    "Skid": "撬块加工 / 认证 / 出口",
    "On Site Engineering Design": "现场工程设计（UK 端）",
    "On Site Installing": "现场安装与开车（UK 端）",
}

# Item-level free-form note seeded into the editable row. Kept empty by
# default so the row备注 stays a clean canvas the user can fill on the page
# (会议结论 / 决议另外在顶部决议面板里）。每个条目都是可编辑的，修改后会
# 保存到本浏览器并可以通过工具栏导出 JSON。
ITEM_NOTES: Dict[str, str] = {}

# Override owner labels for clarity
OWNER_LABEL = {
    "UK": "UK",
    "SW": "UK Sandwich (SW)",
    "IEPE": "IEPE",
    "CFCT": "CFCT",
    "CIMT": "CIMT",
    "IEPE/UK": "IEPE + UK",
    "IEPE/CFCT": "IEPE + CFCT",
    "IEPE/CFCT/UK": "IEPE + CFCT + UK",
    "IEPE/CFCT/CIMT": "IEPE + CFCT + CIMT",
    "CFCT/IEPE/UK": "CFCT + IEPE + UK",
    "CFCT/CIMT": "CFCT + CIMT",
    "UK/CFCT/IEPE": "UK + CFCT + IEPE",
    "UK/IEPE": "UK + IEPE",
    "Import/Export": "进出口部",
}


# ---- decisions / actions extracted from the 2026-05-09 meetings --------------

DECISIONS = [
    {
        "topic": "控制策略：减压阀 + 质量流量控制器 + 液位",
        "context": (
            "气液分离罐压力 10–68 bar，下游产品罐常压；单一调节阀无法降到常压。"
            " 高宇选用 Bronkhorst 质量流量控制器（流量+调节一体，内置 PID）。"
        ),
        "decision": (
            "保留 1 个减压阀 + 流量控制器，在 DCS 上按 \"液位 → 流量控制器\" 实施，"
            " 流量控制器视为执行机构（内置 PID），不再做单独串级。"
            " 减压阀改用 Emerson Fisher 自力式机械减压阀（替代原电动方案，省成本/省空间）。"
        ),
        "owner": "孟德智 起草反馈 + 高宇 更新仪表条件",
    },
    {
        "topic": "防爆分区（Hazardous Area Classification）",
        "context": (
            "UK Keith 给的预判：软口 0.5m + 开口 1m 范围内多处 Zone 1；"
            " 与国内 GB 50058 通常做法（通风良好整体 Zone 2，仅密封点小范围 Zone 1）不一致。"
        ),
        "decision": (
            "我方推进 3D 模型冻结后，由 UK 属地按当地规范出爆炸危险区域图；"
            " 我方根据该图调整：① 移开管口/触摸屏躲开 Zone 1；②无法移动则提高仪表防爆等级；"
            " 质量流量控制器无法做到 Zone 1，必须躲开。"
        ),
        "owner": "胡超群 推进 3D；UK 出 HAC 图",
    },
    {
        "topic": "采购流程 vs UK 选型确认 (Keith)",
        "context": (
            "UK 希望签 PO 前看到 vendor + model；国内采购必须比价 3 家，"
            " 数据单不允许指定品牌。已发生过：泵从 3→1→2→1 反复，采购流程被打回。"
        ),
        "decision": (
            "在 \"采购倾向于某家但未签合同\" 这个时间窗内，由对应专业工程师"
            " (仪表→高宇；设备→赵子亮/范双双) 把该厂家资料发 Keith 审一次；"
            " 在采购流程节点加 \"传阅给 Keith\" 的拦截。"
        ),
        "owner": "赵子亮 落地拦截点；高宇 / 范双双 执行发文。",
    },
    {
        "topic": "HAZOP 与 3D 模型的顺序",
        "context": (
            "UK HAZOP 远比国内详细 (含 What-if、日常操作)，要求 3D 冻结+全文件齐备才做；"
            " 同时 UK 又希望 vendor 选型先行 — 与 ① 矛盾。"
        ),
        "decision": (
            "整体调整：① 5月底/6月初先把 3D 模型 + 设备/安全阀计算 update 完；"
            " ② 6/2-6/3 周做 HAZOP；③ 7月初再发起询价 / 选型 → Keith 审；"
            " ④ 8/1 周签 CE 合同 → 主设备签 PO；④ 供货 ~4 个月。"
            " 控制说明出版稿 6 月底完成（不影响采购，可滞后到 7/1 周）。"
        ),
        "owner": "胡超群（推进3D）+ 李涛博士（与 UK 沟通顺序）",
    },
    {
        "topic": "反应柱 1.5 m 高度争议",
        "context": "UK 反复以为是床层高度，实际为反应柱整体外形高度。",
        "decision": "下周（5/12 周）发出书面说明，附最新规格书与图。",
        "owner": "范双双 出图、赵子亮复核、胡超群对外沟通。",
    },
    {
        "topic": "设备数据单更新 (BT06、反应柱 DN)",
        "context": "BT06 容积变更 220L→20-30L；反应柱 01 改 DN80、02/03 改 DN65；旧数据单未同步。",
        "decision": "5月底前所有 equipment data sheet 与 PI&D 管口对齐 update；下周先出主设备版。",
        "owner": "范双双（CFCT）+ 胡超群（IEPE）。",
    },
    {
        "topic": "概算号 / 项目号",
        "context": (
            "孟德智部门没有概算号，发起采购被卡；项目号尚未申请；"
            " 历史上 FMO/中化学项目走 \"项目组\" 申请。"
        ),
        "decision": (
            "由 CFCT 侧（陶总）统一申请一个项目号（避免分拆）；"
            " 控制系统、电缆桥架等如需拉工程部协助，事后再开专题会。"
        ),
        "owner": "赵子亮",
    },
    {
        "topic": "材质等级（与氢气接触必须 ≥316L）",
        "context": "SGS 早期反馈：H₂ 接触面最低 316L。当前等级表里有 304 用于公用 / 尾气段。",
        "decision": "保留 H₂ 接触段 316L；公用工程 / PSV 后尾气段 304。",
        "owner": "程英辉（材质）+ CFCT 采购。",
    },
]

ACTIONS = [
    ("下周（5/12 周）", "孟德智", "起草控制策略反馈邮件（液位→质量流量控制器 + 自力式减压阀）"),
    ("下周（5/12 周）", "高宇", "更新仪表条件表，将控制阀型号从电动改为机械式 Fisher 自力式"),
    ("下周（5/12 周）", "范双双 / 赵子亮", "出反应柱 / BT06 等设备 data sheet 更新版"),
    ("下周（5/12 周）", "胡超群", "把 1.5m 高度解释 + 反应柱 DN80/DN65 调整正式发 UK"),
    ("下周（5/12 周）", "赵子亮", "把 ATEX / MD / EMC 资料补发 Keith（PED 已发）"),
    ("5月底", "胡超群 + CFCT", "3D 模型冻结草版，发 UK 审查"),
    ("5月底", "范双双", "完成 PSV 计算 + 安全阀数据单"),
    ("6月", "孟德智", "控制说明书出版稿 (出版 6 月底)"),
    ("6/2-6/3 周", "UK + IEPE", "HAZOP（3D 模型基础上）"),
    ("6 月", "高宇", "仪表数据单 update（基于 PI&D 终稿）"),
    ("7/1 周", "范双双 + 高宇", "采购询价文件 + Keith 复核启动"),
    ("8/1 周", "赵子亮", "签 CE 认证合同；主设备 PO 签订（供货 ~4 个月）"),
    ("常态", "赵子亮", "在采购流程加入 \"签合同前传阅给 Keith\" 拦截点"),
    ("待定", "赵子亮 / 高少峰", "申请统一项目号（CFCT 出口路径）"),
]

OPEN_RISKS = [
    {
        "title": "3D 模型是新基线的关键路径",
        "detail": "如 5月底未冻结，HAZOP / 采购 / 防爆分区 / Keith 审都顺延，整体可能再后推 2-4 周。",
        "severity": "高",
    },
    {
        "title": "UK vendor 选型期望 vs 国内 3 家比价流程",
        "detail": "如采购拦截点未落实，会出现 \"采购已下单但 Keith 未审\" 的合规风险。",
        "severity": "高",
    },
    {
        "title": "防爆分区图依赖 UK 属地出图",
        "detail": "若 UK 不能在 6 月内出 HAC 图，则仪表选型、控制柜位置无法定稿。",
        "severity": "中",
    },
    {
        "title": "反应柱 1.5m 反复理解偏差",
        "detail": "若 UK 不接受我方解释，可能引发反应柱重新设计 → data sheet / BOM / 价格全部翻修。",
        "severity": "中",
    },
    {
        "title": "运输/出口周期 3 个月",
        "detail": "需进出口部尽早介入；当前 timeline 中 export shipment 占 2027/3-2027/6。",
        "severity": "中",
    },
    {
        "title": "控制柜及电缆桥架采购归属未明",
        "detail": "孟德智部门无概算号；历史走工程部，但本次工程部尚未介入。",
        "severity": "中",
    },
]

# NOTE: Per 5/11 feedback the Gantt now mirrors the xlsx exactly — no
# extra bars are layered on top of the source data. Meeting-derived
# scheduling lives in the decisions / actions / risks panels above the
# chart, not as fake bars.

# Dependency map: deliverable -> list of upstream deliverables it depends on.
# Names must match exactly the `name_en` keys used in the xlsx.
DEPENDENCIES = {
    "Equipment datasheets": ["PI&D", "Reactor column specification sheet"],
    "Reactor column specification sheet": ["PI&D"],
    "Preliminary equipment layout": ["PI&D", "Process Equipment List"],
    "Main Safety Relief Facilities Data Sheet": ["PI&D", "Equipment datasheets"],
    "Pump specification sheet": ["PI&D", "Process Equipment List"],
    "Instrument datasheets": ["PI&D", "Instrumentation Condition Table"],
    "Interlock logic": ["PI&D", "Control Description"],
    "3D modelling (procurement feedback; UK confirmation of equipment layout)": [
        "Equipment datasheets",
        "Preliminary equipment layout",
        "Main Safety Relief Facilities Data Sheet",
        "Pump specification sheet",
        "Reactor column specification sheet",
    ],
    "HAZOP （support resource in grey）": [
        "3D modelling (procurement feedback; UK confirmation of equipment layout)",
        "PI&D",
        "Control Description",
        "Equipment datasheets",
    ],
    "Procurement information confirmation(pump\\insturment\\equipment schedule)": [
        "3D modelling (procurement feedback; UK confirmation of equipment layout)",
        "HAZOP （support resource in grey）",
        "Equipment datasheets",
        "Instrument datasheets",
        "Pump specification sheet",
    ],
    "Fabrication drawings": [
        "3D modelling (procurement feedback; UK confirmation of equipment layout)",
    ],
    "BOM": [
        "Fabrication drawings",
        "3D modelling (procurement feedback; UK confirmation of equipment layout)",
    ],
    "Engineering procurement (valves & instruments, static equipment, pumps, piping and fittings; within the skid)": [
        "Procurement information confirmation(pump\\insturment\\equipment schedule)",
        "BOM",
    ],
    "Control system procurement and configuration": [
        "Hardware drawing of PLC",
        "IOA of PLC",
        "Cable schedule",
    ],
    "CE certification (individual equipment + whole-skid certification per Notified Body requirements); full documentation completed": [
        "Procurement information confirmation(pump\\insturment\\equipment schedule)",
        "HAZOP （support resource in grey）",
    ],
    "Equipment fabrication, assembly and debugging": [
        "Engineering procurement (valves & instruments, static equipment, pumps, piping and fittings; within the skid)",
        "Control system procurement and configuration",
        "BOM",
        "Fabrication drawings",
    ],
    "Integrated factory acceptance (FAT)": [
        "Equipment fabrication, assembly and debugging",
        "CE certification (individual equipment + whole-skid certification per Notified Body requirements); full documentation completed",
    ],
    "Export formalities, transport inspection/certification and packing": [
        "Integrated factory acceptance (FAT)",
        "CE certification (individual equipment + whole-skid certification per Notified Body requirements); full documentation completed",
    ],
    "Export shipment (Europe)": [
        "Export formalities, transport inspection/certification and packing",
    ],
    "Skid installation": [
        "Export shipment (Europe)",
        "Off-skid/peripheral installation",
    ],
    "Mechanical, electrical, instrumentation & piping (MEIP) installation": [
        "Skid installation",
    ],
    "Qualification & acceptance testing (DQ / IQ / OQ)": [
        "Mechanical, electrical, instrumentation & piping (MEIP) installation",
    ],
    "Commissioning and start-up": [
        "Qualification & acceptance testing (DQ / IQ / OQ)",
    ],
}

TEAM = [
    ("UK Sandwich", "现场使用方 / 需求方", "Keith 等"),
    ("IEPE（国内）", "整体设计 + 主导对接 UK", "高少峰 (lead) / 胡超群 (执行) / 孟德智 (控制+电缆) / 高宇 (仪表)"),
    ("CFCT（国内）", "撬块 layout / 反应器 / 采购 / 组装 / CE 认证 / 运输",
     "赵子亮 (lead) / 范双双 (执行)"),
    ("CIMT", "PLC 控制系统硬件、IOA、电缆", "—"),
    ("Keith / NB", "CE / ATEX / MD 第三方认证联系人 + Notified Body", "外部"),
    ("进出口部", "出口手续 / 海运 / 清关", "—"),
]

LEGEND = [
    ("original",  "浅蓝", "原计划（旧基线）"),
    ("revised",   "橙黄", "本次更新后的新计划"),
    ("update",    "紫色", "需要更新的时间节点"),
    ("cert",      "绿色", "CE 认证公司介入节点"),
]


# ---- HTML rendering ----------------------------------------------------------

def slugify(name: str) -> str:
    """Stable, short ASCII slug for use as DOM id / localStorage key."""
    import re
    s = re.sub(r"[^a-zA-Z0-9]+", "-", name).strip("-").lower()
    return s[:64] or "item"


def render(items: List[Item], col_start: int, col_end: int) -> str:
    # Build month header info
    month_headers: List[Tuple[int, int, int]] = []  # (year, month, span)
    cur_y, cur_m, _ = col_to_month_week(col_start)
    cur_count = 0
    cur_year_month = (cur_y, cur_m)
    for c in range(col_start, col_end + 1):
        y, m, _ = col_to_month_week(c)
        if (y, m) != cur_year_month and cur_count:
            month_headers.append((cur_year_month[0], cur_year_month[1], cur_count))
            cur_year_month = (y, m)
            cur_count = 0
        cur_count += 1
    if cur_count:
        month_headers.append((cur_year_month[0], cur_year_month[1], cur_count))

    today_col = date_to_col(TODAY)

    n_cols = col_end - col_start + 1
    week_w = 18                      # px per week column
    label_w = 380                    # px for left fixed label column
    today_left_px = label_w + (today_col - col_start) * week_w
    today_left_pct = (today_col - col_start) / (col_end - col_start + 1) * 100.0

    # Group items by section + assign stable keys
    sections: Dict[str, List[Item]] = {}
    for it in items:
        sections.setdefault(it.section or "Other", []).append(it)

    item_key: Dict[str, str] = {}
    used = set()
    for it in items:
        base = slugify(it.name_en)
        k = base
        i = 1
        while k in used:
            i += 1
            k = f"{base}-{i}"
        used.add(k)
        item_key[it.name_en] = k

    # Build rows HTML
    rows_html: List[str] = []
    track_w = n_cols * week_w
    for sec_key, sec_items in sections.items():
        sec_zh = SECTION_LABEL_ZH.get(sec_key, sec_key)
        rows_html.append(
            f'<div class="section-row">'
            f'<div class="section-label">{html.escape(sec_zh)}'
            f'<span class="section-en">{html.escape(sec_key)}</span></div>'
            f'<div class="section-bar" style="width:{track_w}px"></div></div>'
        )
        for it in sec_items:
            zh = ITEM_LABEL_ZH.get(it.name_en, it.name_en)
            owner_label = OWNER_LABEL.get(it.owner, it.owner)
            note = ITEM_NOTES.get(it.name_en, "")
            key = item_key[it.name_en]
            deps = DEPENDENCIES.get(it.name_en, [])
            dep_keys = [item_key[d] for d in deps if d in item_key]

            # Build bars per bucket as contiguous runs of same-bucket cells.
            cells_by_col: Dict[int, Cell] = {}
            for c in it.cells:
                # Prefer non-note buckets when colliding (EXTRA_BARS use non-note buckets)
                ex = cells_by_col.get(c.col)
                if not ex or (ex.bucket == "note" and c.bucket != "note"):
                    cells_by_col[c.col] = c
            bars: List[Tuple[int, int, str, List[str]]] = []
            cur_start = None
            cur_bucket = None
            cur_notes: List[str] = []
            for c in range(col_start, col_end + 2):
                cell = cells_by_col.get(c)
                bucket = cell.bucket if cell else None
                if bucket == "note":
                    bucket = None
                if bucket != cur_bucket:
                    if cur_bucket and cur_start is not None:
                        bars.append((cur_start, c - 1, cur_bucket, cur_notes))
                    cur_bucket = bucket
                    cur_start = c
                    cur_notes = []
                if cell and cell.note:
                    cur_notes.append(cell.note)

            bucket_label = {
                "original": "原计划", "revised": "新计划",
                "update": "需更新", "cert": "CE 介入",
            }

            # Build bar elements
            bar_html = []
            first_bar_left = None
            for (s, e, bucket, notes) in bars:
                left = (s - col_start) * week_w
                width = (e - s + 1) * week_w
                if first_bar_left is None:
                    first_bar_left = left
                ys, ms, ws_ = col_to_month_week(s)
                ye, me, we_ = col_to_month_week(e)
                date_label = (
                    f"{ys}/{ms:02d} 第{ws_}周  →  {ye}/{me:02d} 第{we_}周"
                    if (ys, ms, ws_) != (ye, me, we_) else f"{ys}/{ms:02d} 第{ws_}周"
                )
                tip_lines = [
                    f"📌 {zh}",
                    f"⏱ {date_label} · {bucket_label.get(bucket, bucket)}",
                    f"👥 负责：{owner_label}",
                ]
                if deps:
                    dep_zh = [ITEM_LABEL_ZH.get(d, d) for d in deps]
                    tip_lines.append("↳ 依赖：" + " / ".join(dep_zh))
                if notes:
                    tip_lines.append("📝 " + " ".join(notes))
                if note:
                    tip_lines.append("💬 " + note)
                tip = "&#10;".join(html.escape(t) for t in tip_lines)
                color = PALETTE.get(bucket, "#888")
                bar_html.append(
                    f'<div class="bar bar-{bucket}" '
                    f'style="left:{left}px;width:{width}px;background:{color}" '
                    f'title="{tip}" data-bucket="{bucket_label.get(bucket, "")}" '
                    f'data-item="{key}"></div>'
                )

            note_html = (
                f'<div class="row-note" contenteditable="true" '
                f'data-item="{key}" data-default="{html.escape(note)}" '
                f'data-placeholder="（点此添加备注/解释，将自动保存到本浏览器）">'
                f'{html.escape(note)}</div>'
            )

            first_bar_attr = (
                f' data-first-bar="{first_bar_left}"' if first_bar_left is not None else ""
            )
            # deps / influence containers are filled at page-load by JS so
            # user edits show consistently across "影响" (reverse) lines.
            rows_html.append(
                f'<div class="item-row" id="row-{key}" data-item="{key}"'
                f' data-default-deps="{html.escape(",".join(dep_keys))}"{first_bar_attr}>'
                f'  <div class="item-label" data-item="{key}" title="点击空白处：滚动到该项的时间段并高亮其依赖项">'
                f'    <div class="item-zh">{html.escape(zh)}</div>'
                f'    <div class="item-en">{html.escape(it.name_en)}</div>'
                f'    <div class="item-owner">负责：{html.escape(owner_label)}</div>'
                f'    <div class="row-deps">'
                f'      <span class="rel-label dep-label">↳ 依赖：</span>'
                f'      <span class="rel-list deps-list" data-kind="deps"></span>'
                f'      <button class="rel-edit-btn" data-kind="deps" title="编辑依赖">✏️</button>'
                f'    </div>'
                f'    <div class="row-influence">'
                f'      <span class="rel-label inf-label">↗ 影响：</span>'
                f'      <span class="rel-list inf-list"></span>'
                f'    </div>'
                f'    {note_html}'
                f'  </div>'
                f'  <div class="item-track" style="width:{track_w}px" data-item="{key}">'
                f'    {"".join(bar_html)}'
                f'  </div>'
                f'</div>'
            )

    # Month header HTML (no pseudo ::before any more — we use a real sticky spacer)
    month_header_html = []
    week_header_html = []
    for (y, m, span) in month_headers:
        cls = "mh-cur" if (y, m) == (TODAY.year, TODAY.month) else ""
        month_header_html.append(
            f'<div class="month-cell {cls}" style="width:{span*week_w}px">'
            f'{y}/<span class="m">{m:02d}</span></div>'
        )
        for w in range(1, span + 1):
            week_header_html.append(
                f'<div class="week-cell" style="width:{week_w}px">w{w}</div>'
            )

    # Decisions
    decisions_html = []
    for d in DECISIONS:
        decisions_html.append(
            f'<details class="decision"><summary>'
            f'<span class="d-topic">{html.escape(d["topic"])}</span></summary>'
            f'<div class="d-body">'
            f'  <div><span class="lbl">背景</span>{html.escape(d["context"])}</div>'
            f'  <div><span class="lbl">结论</span>{html.escape(d["decision"])}</div>'
            f'  <div><span class="lbl">负责</span>{html.escape(d["owner"])}</div>'
            f'</div></details>'
        )

    actions_html = []
    for (when, who, what) in ACTIONS:
        actions_html.append(
            f'<tr><td class="when">{html.escape(when)}</td>'
            f'<td class="who">{html.escape(who)}</td>'
            f'<td>{html.escape(what)}</td></tr>'
        )

    risks_html = []
    for r in OPEN_RISKS:
        sev_cls = {"高": "sev-high", "中": "sev-mid", "低": "sev-low"}.get(r["severity"], "")
        risks_html.append(
            f'<div class="risk {sev_cls}">'
            f'  <div class="risk-head"><span class="sev">{r["severity"]}风险</span>'
            f'  {html.escape(r["title"])}</div>'
            f'  <div class="risk-body">{html.escape(r["detail"])}</div>'
            f'</div>'
        )

    legend_html = []
    for (lkey, color_zh, desc) in LEGEND:
        legend_html.append(
            f'<span class="leg"><span class="leg-swatch" '
            f'style="background:{PALETTE[lkey]}"></span>{color_zh} — {desc}</span>'
        )
    minimap_legend_html = []
    for (lkey, color_zh, desc) in LEGEND:
        minimap_legend_html.append(
            f'<span><i style="background:{PALETTE[lkey]}"></i>{color_zh}</span>'
        )

    # Serialize per-item metadata so the JS can build pickers/influence maps.
    items_meta = []
    default_deps_map: Dict[str, List[str]] = {}
    for it in items:
        k = item_key[it.name_en]
        items_meta.append({
            "key": k,
            "zh":  ITEM_LABEL_ZH.get(it.name_en, it.name_en),
            "en":  it.name_en,
            "section": SECTION_LABEL_ZH.get(it.section or "", it.section or ""),
        })
        default_deps_map[k] = [
            item_key[d] for d in DEPENDENCIES.get(it.name_en, []) if d in item_key
        ]
    items_meta_json = json.dumps(items_meta, ensure_ascii=False)
    default_deps_json = json.dumps(default_deps_map, ensure_ascii=False)

    css = """
    :root { --label-w: %(label_w)dpx; --week-w: %(week_w)dpx; }
    * { box-sizing: border-box; }
    body { font-family: 'Helvetica Neue', 'PingFang SC', 'Microsoft YaHei', sans-serif;
           margin: 0; color: #1f2937; background: #f5f7fa; }
    h2 { font-size: 16px; margin: 24px 0 8px 0; color: #0f3460; }
    .container { max-width: 1480px; margin: 0 auto; padding: 24px; }
    .panel { background: #fff; border-radius: 10px; padding: 16px;
             box-shadow: 0 1px 3px rgba(15,52,96,0.06); margin-bottom: 16px; }
    .legend { display: flex; gap: 18px; flex-wrap: wrap; font-size: 13px;
              color: #475569; align-items: center; }
    .leg { display: flex; align-items: center; gap: 6px; }
    .leg-swatch { width: 14px; height: 14px; border-radius: 3px;
                  display: inline-block; }
    table.simple { width: 100%%; border-collapse: collapse; font-size: 13px; }
    table.simple td, table.simple th { padding: 6px 10px; border-bottom: 1px solid #e2e8f0;
                                        vertical-align: top; text-align: left; }
    table.simple th { background: #eef2f7; font-weight: 600; }
    table.simple .when { white-space: nowrap; color: #d97706; font-weight: 600; }
    table.simple .who { white-space: nowrap; color: #0f766e; }
    .decision { background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 6px;
                margin-bottom: 8px; }
    .decision summary { padding: 8px 12px; cursor: pointer; font-weight: 600;
                        color: #0f3460; }
    .decision summary::-webkit-details-marker { color: #d97706; }
    .d-body { padding: 4px 14px 12px 14px; font-size: 13px; line-height: 1.6; }
    .d-body .lbl { display: inline-block; min-width: 38px; color: #d97706;
                   font-weight: 600; margin-right: 6px; }
    .d-body div { margin: 4px 0; }
    .risk { background: #fff7ed; border-left: 3px solid #f59e0b;
            padding: 8px 12px; border-radius: 4px; margin-bottom: 6px; font-size: 13px; }
    .risk.sev-high { background: #fef2f2; border-left-color: #dc2626; }
    .risk.sev-mid  { background: #fff7ed; border-left-color: #f59e0b; }
    .risk.sev-low  { background: #f0fdf4; border-left-color: #16a34a; }
    .risk-head .sev { background: #dc2626; color: #fff; padding: 1px 6px;
                      border-radius: 3px; font-size: 11px; margin-right: 6px; }
    .risk.sev-mid  .sev { background: #f59e0b; }
    .risk.sev-low  .sev { background: #16a34a; }
    .risk-body { color: #475569; margin-top: 2px; }

    /* ---- Gantt panel ---- */
    .gantt-toolbar { display: flex; gap: 14px; flex-wrap: wrap;
                     align-items: center; justify-content: space-between;
                     padding: 6px 8px 12px 8px; border-bottom: 1px solid #e2e8f0;
                     margin-bottom: 10px; }
    .gantt-toolbar .left  { display: flex; gap: 18px; flex-wrap: wrap;
                            align-items: center; font-size: 13px; color: #475569; }
    .gantt-toolbar .right { display: flex; gap: 8px; }
    .toolbar-btn { background: #0f3460; color: #fff; border: none;
                   padding: 6px 12px; border-radius: 6px; cursor: pointer;
                   font-size: 12px; }
    .toolbar-btn:hover { background: #16213e; }
    .toolbar-btn.secondary { background: #fff; color: #0f3460;
                              border: 1px solid #cbd5e1; }
    .toolbar-btn.secondary:hover { background: #f1f5f9; }
    .toolbar-hint { font-size: 11px; color: #94a3b8; margin-top: 4px; }

    .gantt { background: #fff; border-radius: 10px; overflow: hidden;
             box-shadow: 0 1px 3px rgba(15,52,96,0.06); }
    .gantt-scroll { overflow: auto; max-height: 78vh; position: relative; }
    .gantt-grid { position: relative; }

    /* sticky-top headers; left-sticky corner uses a real spacer div */
    .month-row, .week-row { display: flex; position: sticky; z-index: 8;
                            background: #f1f5f9; }
    .month-row { top: 0; font-weight: 600; color: #0f3460; }
    .week-row  { top: 28px; color: #64748b; font-size: 11px; }
    .header-spacer { width: var(--label-w); flex-shrink: 0;
                      background: #e2e8f0; border-right: 1px solid #cbd5e1;
                      position: sticky; left: 0; z-index: 10; }
    .month-cell { padding: 6px 8px; border-right: 1px solid #cbd5e1;
                  text-align: center; flex-shrink: 0; }
    .month-cell.mh-cur { background: #fde68a; color: #92400e; }
    .month-cell .m { font-size: 13px; }
    .week-cell { border-right: 1px dashed #e2e8f0; padding: 2px 0;
                 text-align: center; flex-shrink: 0; }

    /* sticky-left labels */
    .section-row { display: flex; align-items: stretch; background: #eef2f7; }
    .section-label { width: var(--label-w); padding: 6px 10px;
                     font-weight: 600; color: #0f3460; flex-shrink: 0;
                     border-right: 1px solid #cbd5e1;
                     position: sticky; left: 0; z-index: 7;
                     background: #eef2f7; }
    .section-label .section-en { color: #94a3b8; font-weight: 400;
                                  font-size: 11px; margin-left: 6px; }
    .section-bar { flex-shrink: 0; }

    .item-row { display: flex; border-bottom: 1px solid #f1f5f9;
                min-height: 40px; position: relative; }
    .item-row:hover .item-label { background: #f1f5f9; }
    .item-row.row-selected { background: #fff7ed; }
    .item-row.row-selected .item-label { background: #ffedd5;
                                          box-shadow: inset 3px 0 0 #d97706; }
    .item-row.row-dep { background: #fef9c3; }
    .item-row.row-dep .item-label { background: #fef08a;
                                     box-shadow: inset 3px 0 0 #facc15; }
    .item-label { width: var(--label-w); padding: 8px 10px;
                  border-right: 1px solid #e2e8f0; flex-shrink: 0;
                  position: sticky; left: 0; z-index: 6;
                  background: #fff; cursor: pointer;
                  transition: background 0.15s ease; }
    .item-zh { font-size: 13px; font-weight: 600; color: #0f172a; }
    .item-en { font-size: 11px; color: #94a3b8; margin-top: 2px;
               word-break: break-word; }
    .item-owner { font-size: 11px; color: #0f766e; margin-top: 2px; }
    .row-deps { font-size: 11px; color: #6b7280; margin-top: 3px;
                line-height: 1.5; }
    .row-deps .dep-link { color: #0369a1; cursor: pointer;
                           text-decoration: underline dotted;
                           text-underline-offset: 2px; }
    .row-deps .dep-link:hover { color: #075985; }
    .row-note { font-size: 11px; color: #475569; background: #fefce8;
                padding: 4px 6px; border-left: 2px solid #facc15;
                margin-top: 4px; line-height: 1.55;
                border-radius: 0 4px 4px 0; outline: none;
                white-space: pre-wrap; min-height: 20px; }
    .row-note:focus { background: #fef3c7; border-left-color: #d97706;
                      box-shadow: 0 0 0 2px rgba(217,119,6,0.15); }
    .row-note.edited { background: #fef3c7; border-left-color: #ea580c; }
    .row-note:empty::before { content: attr(data-placeholder);
                               color: #cbd5e1; font-style: italic; }

    .item-track { position: relative; flex-shrink: 0; }
    .bar { position: absolute; top: 8px; height: 22px; border-radius: 3px;
           opacity: 0.92; cursor: help;
           transition: transform 0.1s ease, box-shadow 0.1s ease; }
    .bar:hover { transform: translateY(-1px);
                  box-shadow: 0 4px 8px rgba(0,0,0,0.15); z-index: 3; }
    .bar.bar-revised  { box-shadow: 0 0 0 1px rgba(0,0,0,0.08); }
    .bar.bar-original { opacity: 0.5; height: 7px; top: 29px; border-radius: 2px; }
    .bar.bar-update   { width: 7px !important; height: 38px !important; top:0;
                          border-radius: 2px; box-shadow: 0 0 0 1px #4c1d95; }
    .bar.bar-cert     { box-shadow: 0 0 0 1px #166534 inset; }
    .row-selected .bar.bar-revised, .row-selected .bar.bar-update,
    .row-selected .bar.bar-cert {
        box-shadow: 0 0 0 2px #d97706; transform: translateY(-1px);
    }

    .today-marker {
       position: absolute; top: 0; bottom: 0; width: 2px; background: #dc2626;
       left: %(today_left_px).1fpx;
       z-index: 4; pointer-events: none;
    }
    .today-label { position: absolute; top: 30px; transform: translateX(-50%%);
                   background: #dc2626; color: #fff; font-size: 11px;
                   padding: 1px 6px; border-radius: 3px; white-space: nowrap; }

    .toast { position: fixed; bottom: 20px; left: 50%%;
             transform: translateX(-50%%) translateY(20px);
             background: #0f3460; color: #fff; padding: 8px 16px;
             border-radius: 6px; font-size: 13px; opacity: 0;
             transition: opacity 0.2s ease, transform 0.2s ease;
             pointer-events: none; z-index: 1000; }
    .toast.show { opacity: 1; transform: translateX(-50%%) translateY(0); }

    /* ---- deps / influence chips ---- */
    .row-deps, .row-influence { font-size: 11px; color: #6b7280;
                                 margin-top: 3px; line-height: 1.7;
                                 display: flex; flex-wrap: wrap; align-items: center;
                                 gap: 4px; }
    .rel-label { color: #6b7280; flex-shrink: 0; }
    .dep-label { color: #0369a1; }
    .inf-label { color: #be185d; }
    .rel-list  { display: inline-flex; flex-wrap: wrap; gap: 4px; flex: 1; }
    .chip { display: inline-flex; align-items: center; gap: 2px;
            background: #e0f2fe; color: #075985; padding: 1px 7px;
            border-radius: 9px; cursor: pointer; font-size: 11px;
            border: 1px solid transparent; }
    .chip:hover { background: #bae6fd; border-color: #0284c7; }
    .chip-inf { background: #fce7f3; color: #9d174d; }
    .chip-inf:hover { background: #fbcfe8; border-color: #db2777; }
    .chip-empty { color: #cbd5e1; font-style: italic; cursor: default; }
    .chip-empty:hover { background: transparent; border-color: transparent; }
    .chip-remove { color: #94a3b8; font-size: 13px; line-height: 1;
                   padding: 0 0 0 2px; }
    .chip-remove:hover { color: #dc2626; }
    .rel-edit-btn { background: none; border: 1px dashed #cbd5e1;
                    border-radius: 9px; padding: 1px 6px; cursor: pointer;
                    font-size: 11px; color: #64748b; }
    .rel-edit-btn:hover { background: #f1f5f9; border-color: #94a3b8;
                          color: #0f3460; }
    .rel-edit-btn.editing { background: #fef3c7; border-color: #d97706;
                            color: #92400e; }

    /* ---- dependency editor popover ---- */
    .dep-editor { position: absolute; z-index: 50; background: #fff;
                  border: 1px solid #cbd5e1; border-radius: 8px;
                  box-shadow: 0 8px 24px rgba(15,52,96,0.18);
                  padding: 10px 12px; width: 360px; font-size: 12px;
                  color: #1f2937; }
    .dep-editor .de-title { font-weight: 600; color: #0f3460;
                            margin-bottom: 6px; font-size: 12px; }
    .dep-editor .de-current { display: flex; flex-wrap: wrap; gap: 4px;
                              min-height: 22px; margin-bottom: 8px;
                              padding-bottom: 6px;
                              border-bottom: 1px dashed #e2e8f0; }
    .dep-editor .de-add-row { display: flex; gap: 6px; }
    .dep-editor select { flex: 1; padding: 4px 6px; border: 1px solid #cbd5e1;
                         border-radius: 4px; font-size: 12px; min-width: 0; }
    .dep-editor .de-actions { display: flex; justify-content: space-between;
                              margin-top: 8px; gap: 6px; }
    .dep-editor button.de-btn { background: #0f3460; color: #fff;
                                 border: none; padding: 4px 10px;
                                 border-radius: 4px; cursor: pointer;
                                 font-size: 12px; }
    .dep-editor button.de-btn.secondary { background: #fff; color: #0f3460;
                                           border: 1px solid #cbd5e1; }
    .dep-editor button.de-btn.danger    { background: #fff; color: #b91c1c;
                                           border: 1px solid #fecaca; }
    .dep-editor button.de-btn:hover { opacity: 0.9; }
    .dep-editor .de-empty { color: #94a3b8; font-style: italic; font-size: 11px; }

    /* ---- minimap (draggable overview) ---- */
    .minimap { position: fixed; right: 20px; bottom: 20px;
               width: 360px; background: #fff;
               border: 1px solid #cbd5e1; border-radius: 8px;
               box-shadow: 0 10px 30px rgba(15,52,96,0.20);
               z-index: 200; user-select: none;
               transition: opacity 0.2s ease; }
    .minimap.collapsed .mm-body { display: none; }
    .minimap .mm-header { display: flex; align-items: center;
                           justify-content: space-between;
                           background: #0f3460; color: #fff;
                           padding: 6px 10px;
                           border-radius: 8px 8px 0 0;
                           cursor: grab; font-size: 12px; }
    .minimap.dragging .mm-header { cursor: grabbing; }
    .minimap.collapsed .mm-header { border-radius: 8px; }
    .minimap .mm-title { display: flex; align-items: center; gap: 6px;
                          font-weight: 600; }
    .minimap .mm-icons { display: flex; gap: 4px; }
    .minimap .mm-icons button {
      background: rgba(255,255,255,0.15); color: #fff; border: none;
      padding: 2px 8px; border-radius: 4px; cursor: pointer;
      font-size: 11px;
    }
    .minimap .mm-icons button:hover { background: rgba(255,255,255,0.3); }
    .minimap .mm-body { padding: 8px; }
    .minimap .mm-canvas { position: relative; background: #f8fafc;
                           border: 1px solid #e2e8f0; border-radius: 4px;
                           overflow: hidden; }
    .minimap .mm-canvas svg { display: block; width: 100%%; height: 100%%; }
    .minimap .mm-viewport { position: absolute; border: 1.5px solid #dc2626;
                             background: rgba(220,38,38,0.10);
                             pointer-events: none; }
    .minimap .mm-canvas.draggable { cursor: crosshair; }
    .minimap .mm-legend { display: flex; flex-wrap: wrap; gap: 8px;
                           font-size: 10px; color: #475569; margin-top: 6px; }
    .minimap .mm-legend span { display: inline-flex; align-items: center; gap: 3px; }
    .minimap .mm-legend i { display: inline-block; width: 9px; height: 9px;
                             border-radius: 2px; }

    .footer { font-size: 11px; color: #94a3b8; text-align: center;
              margin: 18px 0; }
    """ % dict(label_w=label_w, week_w=week_w, today_left_px=today_left_px)

    html_doc = f"""<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8" />
<title>Sandwich 连续氢化项目 — 时间表与 5/9 会议跟踪</title>
<style>{css}</style>
</head>
<body>
<div class="container">

<div class="panel">
  <h2 style="margin-top:0">5/9 会议关键决议（点击展开）</h2>
  {"".join(decisions_html)}
</div>

<div class="panel">
  <h2>下一步行动 (Action items)</h2>
  <table class="simple">
    <thead><tr><th>时间</th><th>责任人</th><th>事项</th></tr></thead>
    <tbody>{"".join(actions_html)}</tbody>
  </table>
</div>

<div class="panel">
  <h2>开放风险 (Open risks)</h2>
  {"".join(risks_html)}
</div>

<div class="panel" id="ganttPanel">
  <h2 style="margin-top:0">甘特图（按交付项 × 周）</h2>
  <div class="gantt-toolbar">
    <div class="left">
      <div class="legend">{"".join(legend_html)}</div>
    </div>
    <div class="right">
      <button class="toolbar-btn secondary" id="btn-scroll-today">⤴ 跳到"今天"</button>
      <button class="toolbar-btn secondary" id="btn-clear-sel">清除高亮</button>
      <button class="toolbar-btn secondary" id="btn-show-mm">🗺 缩略图</button>
      <button class="toolbar-btn" id="btn-export-notes">⇩ 导出修改</button>
      <button class="toolbar-btn secondary" id="btn-reset-all">重置全部</button>
    </div>
  </div>
  <div class="toolbar-hint">
    每行有上下两条：上方粗条（橙=新计划 / 紫=需更新节点 / 绿=CE 介入节点）是当前基线，下方细的浅蓝是原计划。
    悬停时间段可看完整信息；点击左侧交付项滚动并高亮其依赖项；✏️ 按钮可编辑依赖（"影响"会自动同步重算）；
    黄色备注框可直接编辑（自动保存到本浏览器，可"导出修改"分享给团队）。
  </div>

  <div class="gantt">
    <div class="gantt-scroll" id="ganttScroll">
      <div class="gantt-grid" id="ganttGrid" style="min-width: calc(var(--label-w) + {n_cols*week_w}px);">
        <div class="month-row">
          <div class="header-spacer" style="top:0;"></div>
          {"".join(month_header_html)}
        </div>
        <div class="week-row">
          <div class="header-spacer" style="top:28px;"></div>
          {"".join(week_header_html)}
        </div>
        {"".join(rows_html)}
        <div class="today-marker"><div class="today-label">今天 {TODAY.month}/{TODAY.day}</div></div>
      </div>
    </div>
  </div>
</div>

<!-- Draggable mini-map / overview -->
<div class="minimap" id="minimap">
  <div class="mm-header" id="mmHeader">
    <div class="mm-title">🗺 全景缩略图</div>
    <div class="mm-icons">
      <button id="mmToggleBtn" title="折叠/展开">—</button>
      <button id="mmHideBtn" title="隐藏（可在工具栏重新打开）">×</button>
    </div>
  </div>
  <div class="mm-body">
    <div class="mm-canvas draggable" id="mmCanvas" style="height: 220px;">
      <svg id="mmSvg" preserveAspectRatio="none" xmlns="http://www.w3.org/2000/svg"></svg>
      <div class="mm-viewport" id="mmViewport"></div>
    </div>
    <div class="mm-legend">
      {"".join(minimap_legend_html)}
      <span style="margin-left:auto;color:#94a3b8;">拖标题移动 · 点画布跳转</span>
    </div>
  </div>
</div>

<div id="toast" class="toast"></div>

<script id="items-meta" type="application/json">{items_meta_json}</script>
<script id="default-deps" type="application/json">{default_deps_json}</script>

<div class="footer">
  生成于 {TODAY.isoformat()} ｜ 自动从 xlsx + 会议转写抽取，下次会议或基线变更后请重新运行 <code>scripts/build_timeline_view.py</code> 即可刷新本页。
</div>

<script>
(function() {{
  // ===== Constants ========================================================
  var LS_NOTE = "sandwich-note:";
  var LS_DEPS = "sandwich-deps:";
  var scroller = document.getElementById('ganttScroll');
  var grid     = document.getElementById('ganttGrid');
  var labelW   = parseInt(getComputedStyle(document.documentElement).getPropertyValue('--label-w'), 10) || 380;
  var weekW    = parseInt(getComputedStyle(document.documentElement).getPropertyValue('--week-w'), 10) || 18;
  var todayLeftPx = {today_left_px:.1f};
  var trackW   = {track_w};
  var ITEMS    = JSON.parse(document.getElementById('items-meta').textContent);
  var DEFAULTS = JSON.parse(document.getElementById('default-deps').textContent);
  var itemByKey = {{}}; ITEMS.forEach(function(it) {{ itemByKey[it.key] = it; }});

  // ===== State: per-item effective dependencies ===========================
  var effDeps = {{}};         // key -> array of upstream keys
  var infMap  = {{}};         // key -> array of downstream keys
  function loadDeps() {{
    ITEMS.forEach(function(it) {{
      var saved = null;
      try {{ saved = localStorage.getItem(LS_DEPS + it.key); }} catch(e) {{}}
      if (saved !== null) {{
        try {{ effDeps[it.key] = JSON.parse(saved); }}
        catch(e) {{ effDeps[it.key] = (DEFAULTS[it.key] || []).slice(); }}
      }} else {{
        effDeps[it.key] = (DEFAULTS[it.key] || []).slice();
      }}
    }});
    recomputeInfluence();
  }}
  function recomputeInfluence() {{
    infMap = {{}};
    Object.keys(effDeps).forEach(function(k) {{
      effDeps[k].forEach(function(d) {{
        if (!infMap[d]) infMap[d] = [];
        if (infMap[d].indexOf(k) < 0) infMap[d].push(k);
      }});
    }});
  }}
  function saveDeps(key) {{
    var def = (DEFAULTS[key] || []).slice().sort().join(',');
    var cur = (effDeps[key] || []).slice().sort().join(',');
    try {{
      if (def === cur) localStorage.removeItem(LS_DEPS + key);
      else localStorage.setItem(LS_DEPS + key, JSON.stringify(effDeps[key]));
    }} catch(e) {{}}
  }}

  // ===== Rendering deps/influence chips ===================================
  function chipHTML(key, kind) {{
    var meta = itemByKey[key];
    if (!meta) return '';
    var label = meta.zh;
    var cls = (kind === 'inf' ? 'chip chip-inf' : 'chip');
    return '<span class="' + cls + '" data-target="' + key + '">' +
            label + '</span>';
  }}
  function renderRowRel(rowEl) {{
    var key = rowEl.getAttribute('data-item');
    var depsList = rowEl.querySelector('.deps-list');
    var infList  = rowEl.querySelector('.inf-list');
    var deps = effDeps[key] || [];
    var infs = (infMap[key] || []);
    depsList.innerHTML = deps.length
      ? deps.map(function(k) {{ return chipHTML(k, 'dep'); }}).join('')
      : '<span class="chip chip-empty">无</span>';
    infList.innerHTML = infs.length
      ? infs.map(function(k) {{ return chipHTML(k, 'inf'); }}).join('')
      : '<span class="chip chip-empty">无</span>';
    // Mark edited state
    var defKey = (DEFAULTS[key] || []).slice().sort().join(',');
    var curKey = deps.slice().sort().join(',');
    var btn = rowEl.querySelector('.rel-edit-btn');
    if (btn) btn.classList.toggle('editing', defKey !== curKey);
  }}
  function renderAllRel() {{
    document.querySelectorAll('.item-row').forEach(renderRowRel);
  }}

  // ===== Click-to-jump + highlight deps ===================================
  function clearHighlights() {{
    document.querySelectorAll('.item-row').forEach(function(r) {{
      r.classList.remove('row-selected'); r.classList.remove('row-dep');
    }});
  }}
  function scrollToItem(key) {{
    var row = document.getElementById('row-' + key);
    if (!row) return;
    clearHighlights();
    row.classList.add('row-selected');
    (effDeps[key] || []).forEach(function(d) {{
      var dr = document.getElementById('row-' + d);
      if (dr) dr.classList.add('row-dep');
    }});
    var firstBar = row.getAttribute('data-first-bar');
    if (firstBar !== null && firstBar !== '') {{
      var targetLeft = labelW + parseFloat(firstBar) - scroller.clientWidth * 0.30;
      scroller.scrollTo({{ left: Math.max(0, targetLeft), behavior: 'smooth' }});
    }}
    row.scrollIntoView({{ block: 'nearest', behavior: 'smooth' }});
    updateMinimapViewport();
  }}

  // ===== Notes (contenteditable, localStorage) ============================
  function restoreNotes() {{
    document.querySelectorAll('.row-note').forEach(function(el) {{
      var key = el.getAttribute('data-item');
      var def = el.getAttribute('data-default') || '';
      var saved = null;
      try {{ saved = localStorage.getItem(LS_NOTE + key); }} catch(e) {{}}
      if (saved !== null && saved !== def) {{
        el.textContent = saved;
        el.classList.add('edited');
      }}
    }});
  }}
  document.querySelectorAll('.row-note').forEach(function(el) {{
    var key = el.getAttribute('data-item');
    el.addEventListener('blur', function() {{
      var txt = el.innerText.trim();
      var dflt = (el.getAttribute('data-default') || '').trim();
      try {{
        if (txt === dflt) {{
          localStorage.removeItem(LS_NOTE + key);
          el.classList.remove('edited');
        }} else {{
          localStorage.setItem(LS_NOTE + key, txt);
          el.classList.add('edited');
        }}
      }} catch(e) {{}}
    }});
    el.addEventListener('click', function(ev) {{ ev.stopPropagation(); }});
  }});

  // ===== Dependency editor popover ========================================
  var editorEl = null;
  function closeEditor() {{
    if (editorEl && editorEl.parentNode) editorEl.parentNode.removeChild(editorEl);
    editorEl = null;
    document.querySelectorAll('.rel-edit-btn').forEach(function(b) {{ b.disabled = false; }});
  }}
  function openEditor(rowEl, btn) {{
    closeEditor();
    var key = rowEl.getAttribute('data-item');
    var current = (effDeps[key] || []).slice();
    var meta = itemByKey[key];
    editorEl = document.createElement('div');
    editorEl.className = 'dep-editor';
    editorEl.innerHTML =
      '<div class="de-title">编辑 "' + meta.zh + '" 的依赖项</div>' +
      '<div class="de-current"></div>' +
      '<div class="de-add-row">' +
      '  <select class="de-add-select"><option value="">-- 选择要新增的上游 --</option></select>' +
      '  <button class="de-btn de-add-btn">+ 添加</button>' +
      '</div>' +
      '<div class="de-actions">' +
      '  <button class="de-btn danger de-reset-btn">恢复默认</button>' +
      '  <button class="de-btn secondary de-close-btn">完成</button>' +
      '</div>';
    document.body.appendChild(editorEl);
    btn.disabled = true;
    // Position the popover under the edit button
    var rect = btn.getBoundingClientRect();
    editorEl.style.top  = (window.scrollY + rect.bottom + 6) + 'px';
    var left = window.scrollX + rect.left;
    if (left + 360 > window.innerWidth - 12) left = window.innerWidth - 372;
    editorEl.style.left = Math.max(12, left) + 'px';

    var curBox    = editorEl.querySelector('.de-current');
    var selectEl  = editorEl.querySelector('.de-add-select');
    var addBtn    = editorEl.querySelector('.de-add-btn');
    var closeBtn  = editorEl.querySelector('.de-close-btn');
    var resetBtn  = editorEl.querySelector('.de-reset-btn');

    function refreshEditor() {{
      var cur = effDeps[key] || [];
      curBox.innerHTML = '';
      if (cur.length === 0) {{
        curBox.innerHTML = '<span class="de-empty">（无依赖）</span>';
      }} else {{
        cur.forEach(function(k) {{
          var c = document.createElement('span');
          c.className = 'chip';
          c.innerHTML = (itemByKey[k] ? itemByKey[k].zh : k) +
                         '<span class="chip-remove" title="移除">✕</span>';
          c.querySelector('.chip-remove').addEventListener('click', function() {{
            effDeps[key] = (effDeps[key] || []).filter(function(x) {{ return x !== k; }});
            saveDeps(key);
            recomputeInfluence(); renderAllRel(); refreshEditor();
          }});
          curBox.appendChild(c);
        }});
      }}
      // Populate select with items not currently in cur and not self
      selectEl.innerHTML = '<option value="">-- 选择要新增的上游 --</option>';
      ITEMS.filter(function(it) {{ return it.key !== key && cur.indexOf(it.key) < 0; }})
           .forEach(function(it) {{
             var o = document.createElement('option');
             o.value = it.key; o.textContent = it.zh;
             selectEl.appendChild(o);
           }});
    }}
    refreshEditor();

    addBtn.addEventListener('click', function() {{
      var v = selectEl.value;
      if (!v) return;
      if ((effDeps[key] || []).indexOf(v) >= 0) return;
      effDeps[key] = (effDeps[key] || []).concat([v]);
      saveDeps(key);
      recomputeInfluence(); renderAllRel(); refreshEditor();
    }});
    resetBtn.addEventListener('click', function() {{
      if (!confirm('恢复该项的默认依赖？')) return;
      effDeps[key] = (DEFAULTS[key] || []).slice();
      saveDeps(key);
      recomputeInfluence(); renderAllRel(); refreshEditor();
    }});
    closeBtn.addEventListener('click', closeEditor);
  }}
  // Close editor when clicking outside
  document.addEventListener('mousedown', function(ev) {{
    if (!editorEl) return;
    if (ev.target.closest('.dep-editor')) return;
    if (ev.target.classList && ev.target.classList.contains('rel-edit-btn')) return;
    closeEditor();
  }});

  // ===== Click handlers ===================================================
  document.body.addEventListener('click', function(ev) {{
    var chip = ev.target.closest('.chip[data-target]');
    if (chip) {{
      ev.stopPropagation();
      scrollToItem(chip.getAttribute('data-target'));
      return;
    }}
    var editBtn = ev.target.closest('.rel-edit-btn');
    if (editBtn) {{
      ev.stopPropagation();
      var rowEl = editBtn.closest('.item-row');
      openEditor(rowEl, editBtn);
      return;
    }}
  }});

  document.querySelectorAll('.item-label').forEach(function(lbl) {{
    lbl.addEventListener('click', function(ev) {{
      if (ev.target.closest('.row-note')) return;
      if (ev.target.closest('.chip')) return;
      if (ev.target.closest('.rel-edit-btn')) return;
      scrollToItem(lbl.getAttribute('data-item'));
    }});
  }});

  document.querySelectorAll('.bar').forEach(function(b) {{
    b.addEventListener('click', function(ev) {{
      ev.stopPropagation();
      scrollToItem(b.getAttribute('data-item'));
    }});
  }});

  // ===== Toolbar buttons =================================================
  function toast(msg) {{
    var t = document.getElementById('toast');
    t.textContent = msg;
    t.classList.add('show');
    setTimeout(function() {{ t.classList.remove('show'); }}, 1800);
  }}

  document.getElementById('btn-scroll-today').addEventListener('click', function() {{
    var target = todayLeftPx - scroller.clientWidth * 0.30;
    scroller.scrollTo({{ left: Math.max(0, target), behavior: 'smooth' }});
  }});
  document.getElementById('btn-clear-sel').addEventListener('click', clearHighlights);

  document.getElementById('btn-export-notes').addEventListener('click', function() {{
    var notes = {{}}, deps = {{}};
    try {{
      for (var i = 0; i < localStorage.length; i++) {{
        var k = localStorage.key(i);
        if (!k) continue;
        if (k.indexOf(LS_NOTE) === 0) notes[k.slice(LS_NOTE.length)] = localStorage.getItem(k);
        if (k.indexOf(LS_DEPS) === 0) {{
          try {{ deps[k.slice(LS_DEPS.length)] = JSON.parse(localStorage.getItem(k)); }} catch(e){{}}
        }}
      }}
    }} catch(e) {{}}
    var nCount = Object.keys(notes).length;
    var dCount = Object.keys(deps).length;
    if (nCount === 0 && dCount === 0) {{ toast('当前没有修改'); return; }}
    var blob = new Blob([JSON.stringify({{ notes: notes, deps: deps }}, null, 2)],
                       {{ type: 'application/json' }});
    var url = URL.createObjectURL(blob);
    var a = document.createElement('a');
    a.href = url;
    a.download = 'sandwich-edits-' + new Date().toISOString().slice(0,10) + '.json';
    document.body.appendChild(a); a.click(); document.body.removeChild(a);
    URL.revokeObjectURL(url);
    toast('已导出：备注 ' + nCount + ' 条 / 依赖 ' + dCount + ' 项');
  }});

  document.getElementById('btn-reset-all').addEventListener('click', function() {{
    if (!confirm('确定要清除所有本浏览器中的备注与依赖修改吗？')) return;
    try {{
      var rm = [];
      for (var i = 0; i < localStorage.length; i++) {{
        var k = localStorage.key(i);
        if (k && (k.indexOf(LS_NOTE) === 0 || k.indexOf(LS_DEPS) === 0)) rm.push(k);
      }}
      rm.forEach(function(k) {{ localStorage.removeItem(k); }});
    }} catch(e) {{}}
    document.querySelectorAll('.row-note').forEach(function(el) {{
      el.textContent = el.getAttribute('data-default') || '';
      el.classList.remove('edited');
    }});
    loadDeps(); renderAllRel();
    toast('已重置所有修改');
  }});

  // ===== Minimap (draggable overview) =====================================
  var mm        = document.getElementById('minimap');
  var mmHeader  = document.getElementById('mmHeader');
  var mmCanvas  = document.getElementById('mmCanvas');
  var mmSvg     = document.getElementById('mmSvg');
  var mmViewport= document.getElementById('mmViewport');
  var mmToggleBtn = document.getElementById('mmToggleBtn');
  var mmHideBtn   = document.getElementById('mmHideBtn');
  document.getElementById('btn-show-mm').addEventListener('click', function() {{
    mm.style.display = 'block';
  }});
  mmHideBtn.addEventListener('click', function() {{ mm.style.display = 'none'; }});
  mmToggleBtn.addEventListener('click', function() {{
    mm.classList.toggle('collapsed');
    mmToggleBtn.textContent = mm.classList.contains('collapsed') ? '+' : '—';
  }});

  // Build minimap SVG from current Gantt DOM. We pick the top of each
  // .item-row relative to grid; track-only area starts at x = labelW
  // (which we ignore -> we shift bars left by -labelW).
  function buildMinimap() {{
    var rows = grid.querySelectorAll('.item-row, .section-row');
    var gridRect = grid.getBoundingClientRect();
    var totalH = grid.offsetHeight;
    // Cap minimum/maximum dimensions
    var canvasW = mmCanvas.clientWidth || 340;
    var canvasH = mmCanvas.clientHeight || 220;
    mmSvg.setAttribute('viewBox', '0 0 ' + trackW + ' ' + totalH);
    mmSvg.setAttribute('width',  canvasW);
    mmSvg.setAttribute('height', canvasH);
    var parts = [];
    // month boundaries: 4 weeks
    for (var m = 0; m <= {n_cols}; m += 4) {{
      var x = m * weekW;
      parts.push('<line x1="' + x + '" y1="0" x2="' + x + '" y2="' + totalH +
                 '" stroke="#e2e8f0" stroke-width="3" />');
    }}
    // section rows + item rows
    rows.forEach(function(row) {{
      var top = row.offsetTop;
      var h   = row.offsetHeight;
      if (row.classList.contains('section-row')) {{
        parts.push('<rect x="0" y="' + top + '" width="' + trackW +
                    '" height="' + h + '" fill="#dbeafe" />');
        return;
      }}
      // bars within row
      var bars = row.querySelectorAll('.bar');
      bars.forEach(function(b) {{
        var left = parseFloat(b.style.left) || 0;
        var w    = parseFloat(b.style.width) || 0;
        var bg   = b.style.background || '#888';
        var bH   = h - 6;
        var bY   = top + 3;
        if (b.classList.contains('bar-original')) {{
          bH = 4; bY = top + h - 8;
        }}
        if (b.classList.contains('bar-update')) {{
          w = Math.max(w, 6);
        }}
        parts.push('<rect x="' + left + '" y="' + bY + '" width="' + w +
                    '" height="' + bH + '" fill="' + bg + '" opacity="0.85" />');
      }});
    }});
    // today line
    parts.push('<line x1="' + (todayLeftPx - labelW) + '" y1="0" x2="' +
               (todayLeftPx - labelW) + '" y2="' + totalH +
               '" stroke="#dc2626" stroke-width="6" />');
    mmSvg.innerHTML = parts.join('');
    updateMinimapViewport();
  }}

  function updateMinimapViewport() {{
    var canvasW = mmCanvas.clientWidth || 340;
    var canvasH = mmCanvas.clientHeight || 220;
    var scaleX = canvasW / trackW;
    var scaleY = canvasH / (grid.offsetHeight || 1);
    var vpLeft = Math.max(0, scroller.scrollLeft - labelW) * scaleX;
    var vpTop  = scroller.scrollTop * scaleY;
    var vpW    = (scroller.clientWidth - labelW) * scaleX;
    var vpH    = scroller.clientHeight * scaleY;
    mmViewport.style.left   = vpLeft + 'px';
    mmViewport.style.top    = vpTop  + 'px';
    mmViewport.style.width  = Math.max(20, vpW) + 'px';
    mmViewport.style.height = Math.max(20, vpH) + 'px';
  }}

  function scrollFromMinimap(ev) {{
    var rect = mmCanvas.getBoundingClientRect();
    var x = ev.clientX - rect.left;
    var y = ev.clientY - rect.top;
    var canvasW = mmCanvas.clientWidth || 340;
    var canvasH = mmCanvas.clientHeight || 220;
    var scaleX = canvasW / trackW;
    var scaleY = canvasH / (grid.offsetHeight || 1);
    var targetLeft = x / scaleX + labelW - scroller.clientWidth / 2;
    var targetTop  = y / scaleY - scroller.clientHeight / 2;
    scroller.scrollTo({{ left: Math.max(0, targetLeft),
                        top: Math.max(0, targetTop),
                        behavior: 'auto' }});
  }}
  var mmDragging = false;
  mmCanvas.addEventListener('mousedown', function(ev) {{
    mmDragging = true; scrollFromMinimap(ev); ev.preventDefault();
  }});
  document.addEventListener('mousemove', function(ev) {{
    if (mmDragging) scrollFromMinimap(ev);
  }});
  document.addEventListener('mouseup', function() {{ mmDragging = false; }});

  scroller.addEventListener('scroll', updateMinimapViewport);
  window.addEventListener('resize', function() {{
    buildMinimap();
  }});

  // ----- Drag the minimap panel itself ------------------------------------
  (function() {{
    var dragX = 0, dragY = 0, startL = 0, startT = 0, dragging = false;
    mmHeader.addEventListener('mousedown', function(ev) {{
      if (ev.target.tagName === 'BUTTON') return;
      dragging = true;
      mm.classList.add('dragging');
      dragX = ev.clientX; dragY = ev.clientY;
      var r = mm.getBoundingClientRect();
      startL = r.left; startT = r.top;
      mm.style.right = 'auto'; mm.style.bottom = 'auto';
      mm.style.left = startL + 'px'; mm.style.top = startT + 'px';
      ev.preventDefault();
    }});
    document.addEventListener('mousemove', function(ev) {{
      if (!dragging) return;
      var nx = startL + (ev.clientX - dragX);
      var ny = startT + (ev.clientY - dragY);
      nx = Math.max(0, Math.min(window.innerWidth - mm.offsetWidth, nx));
      ny = Math.max(0, Math.min(window.innerHeight - 40, ny));
      mm.style.left = nx + 'px';
      mm.style.top  = ny + 'px';
    }});
    document.addEventListener('mouseup', function() {{
      dragging = false;
      mm.classList.remove('dragging');
    }});
  }})();

  // ===== Init =============================================================
  loadDeps();
  renderAllRel();
  restoreNotes();
  buildMinimap();
  (function() {{
    var target = todayLeftPx - scroller.clientWidth * 0.30;
    scroller.scrollLeft = Math.max(0, target);
    updateMinimapViewport();
  }})();
}})();
</script>

</div>
</body></html>
"""
    return html_doc


def main() -> None:
    items, c0, c1 = load_items()
    out = render(items, c0, c1)
    with open(HTML_PATH, "w", encoding="utf-8") as fh:
        fh.write(out)
    print(f"wrote {HTML_PATH} ({len(out)} bytes, {len(items)} rows)")


if __name__ == "__main__":
    main()
