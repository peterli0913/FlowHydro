"""
Build an internal HAZOP template as an Excel workbook.

Structure:
  Sheet 1 — Cover (study identification)
  Sheet 2 — Node List (system decomposition)
  Sheet 3 — HAZOP Worksheet (main deliverable, pre-populated deviations)
  Sheet 4 — Guide Word / Parameter Reference
  Sheet 5 — Risk Matrix
  Sheet 6 — Recommendations Log
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

# ============ Styling helpers ============
NAVY_FILL   = PatternFill("solid", fgColor="0F3460")
NAVY_LIGHT  = PatternFill("solid", fgColor="DBE4F0")
ORANGE_FILL = PatternFill("solid", fgColor="D97706")
GREEN_FILL  = PatternFill("solid", fgColor="16A34A")
RED_FILL    = PatternFill("solid", fgColor="DC2626")
YELLOW_FILL = PatternFill("solid", fgColor="FEF3C7")
LIGHT_GRAY  = PatternFill("solid", fgColor="F1F5F9")
GREEN_LIGHT = PatternFill("solid", fgColor="DCFCE7")
RED_LIGHT   = PatternFill("solid", fgColor="FEE2E2")
ORANGE_LIGHT= PatternFill("solid", fgColor="FED7AA")

WHITE_BOLD  = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
NAVY_BOLD   = Font(name="Microsoft YaHei", size=11, bold=True, color="0F3460")
BODY_FONT   = Font(name="Microsoft YaHei", size=10)
BODY_BOLD   = Font(name="Microsoft YaHei", size=10, bold=True)
SMALL_FONT  = Font(name="Microsoft YaHei", size=9)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN = Side(style="thin", color="BFC4CC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def header_row(ws, row, headers, fill=NAVY_FILL, font=WHITE_BOLD, height=32):
    ws.row_dimensions[row].height = height
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = fill
        c.font = font
        c.alignment = CENTER
        c.border = BORDER

def data_row(ws, row, values, fill=None, wrap=True, height=None):
    if height:
        ws.row_dimensions[row].height = height
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BODY_FONT
        c.alignment = LEFT_TOP if wrap else LEFT
        c.border = BORDER
        if fill:
            c.fill = fill


wb = Workbook()

# ==================================================================
# Sheet 1: Cover 封面
# ==================================================================
ws = wb.active
ws.title = "1. 封面 Cover"

set_col_widths(ws, [4, 30, 60, 20])
ws.row_dimensions[1].height = 42

ws.merge_cells("A1:D1")
c = ws.cell(row=1, column=1, value="内部 HAZOP 分析工作表 ｜ Internal HAZOP Study Worksheet")
c.fill = NAVY_FILL
c.font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
c.alignment = CENTER

ws.merge_cells("A2:D2")
c = ws.cell(row=2, column=1,
            value="SW Sandwich 连续氢化撬块 · Pre-HAZOP internal review · 2026")
c.fill = NAVY_LIGHT
c.font = NAVY_BOLD
c.alignment = CENTER
ws.row_dimensions[2].height = 24

cover_rows = [
    ("A. 研究基本信息 ｜ Study identification", None, None),
    ("A1", "项目名称 Project name", "SW Sandwich Continuous Hydrogenation Skid ｜ 凯莱英 UK Sandwich 连续氢化撬块"),
    ("A2", "分析日期 Study date", "____ / ____ / 2026"),
    ("A3", "分析地点 Location", "IEPE 会议室 / Online"),
    ("A4", "分析类型 Study type", "内部 HAZOP（正式 HAZOP 前的内部预分析）"),
    ("A5", "分析范围 Scope",
        "从原料储罐 BT01/02 进料，经计量泵 MP01/02、预热器、连续反应器 CR01/CR02/CR03、"
        "冷凝器、气液分离器 SE01，至产品收集罐 BT04；含 H2 供给、放空 / 泄放系统、"
        "公用工程接口、清洗系统、取样系统与控制系统"),
    ("A6", "分析目的 Study objective",
        "在 UK 方正式 HAZOP 前，由国内团队开展内部预分析，识别设计缺陷与遗留问题；"
        "输出建议措施清单，作为后续正式 HAZOP 前的设计闭环依据"),
    ("", "", ""),
    ("B. 分析小组 ｜ Study team", None, None),
    ("B1", "主席 Chair", "___________"),
    ("B2", "工艺 Process",  "邵峰 / 胡超群"),
    ("B3", "静设备 Static Equipment", "赵子亮 / 范双双"),
    ("B4", "控制 Control",  "孟德智"),
    ("B5", "仪表 Instrumentation", "高宇"),
    ("B6", "材料 Materials", "陈辉"),
    ("B7", "记录 Recorder", "Peter"),
    ("B8", "其它 Others", "李涛博士 / Alex 博士（如可以出席）"),
    ("", "", ""),
    ("C. 参考文件 ｜ Reference documents", None, None),
    ("C1", "PFD",            "Revision ____ / 日期 ____"),
    ("C2", "P&ID",           "Revision ____ / 日期 ____"),
    ("C3", "URS",            "Revision ____ / 日期 ____"),
    ("C4", "设备数据单 Equipment Datasheet", "Revision ____ / 日期 ____"),
    ("C5", "仪表数据单 Instrument Datasheet", "Revision ____ / 日期 ____"),
    ("C6", "控制策略 Control Philosophy", "Revision ____ / 日期 ____"),
    ("C7", "泄放设计研究 Relief Design Study", "Revision ____ / 日期 ____"),
    ("C8", "3D 模型 3D Model", "Revision ____ / 日期 ____"),
    ("", "", ""),
    ("D. 使用说明 ｜ Instructions", None, None),
    ("D1", "步骤 1", "对照 Sheet 2『节点清单』，逐节点分析"),
    ("D2", "步骤 2", "对每个节点，参考 Sheet 4 的引导词×参数矩阵，识别偏离"),
    ("D3", "步骤 3", "对每个偏离，识别原因、后果、已有防护，进行风险评级"),
    ("D4", "步骤 4", "对风险不可接受的偏离，输出建议措施到 Sheet 6"),
    ("D5", "步骤 5", "所有建议措施须指定负责人与截止日期"),
]

row_ptr = 4
for tag, key, val in cover_rows:
    if key is None:
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=4)
        c = ws.cell(row=row_ptr, column=1, value=tag)
        c.fill = ORANGE_FILL
        c.font = WHITE_BOLD
        c.alignment = LEFT
        ws.row_dimensions[row_ptr].height = 22
    else:
        ws.cell(row=row_ptr, column=1, value=tag).font = SMALL_FONT
        ws.cell(row=row_ptr, column=2, value=key).font = BODY_BOLD
        ws.cell(row=row_ptr, column=2).fill = LIGHT_GRAY
        c = ws.cell(row=row_ptr, column=3, value=val)
        c.font = BODY_FONT
        c.alignment = LEFT_TOP
        ws.merge_cells(start_row=row_ptr, start_column=3, end_row=row_ptr, end_column=4)
        for col in range(1, 5):
            ws.cell(row=row_ptr, column=col).border = BORDER
        ws.row_dimensions[row_ptr].height = 30 if len(val or "") < 60 else 55
    row_ptr += 1


# ==================================================================
# Sheet 2: Node List 节点清单
# ==================================================================
ws2 = wb.create_sheet("2. 节点清单 Node List")
set_col_widths(ws2, [8, 30, 50, 30, 15])

header_row(ws2, 1, ["节点号\nNode#", "节点名称\nNode description",
                    "设计意图\nDesign intent",
                    "P&ID 参考\nP&ID ref",
                    "关键设备\nKey equipment"])

nodes = [
    ("N1", "原料储罐与计量泵 ｜\nFeed tanks & metering pumps",
     "储存反应液原料并按工艺流量向反应器组稳定进料",
     "PID-XXX Zone A",
     "BT01, BT02, MP01, MP02"),
    ("N2", "H₂ 供给系统 ｜\nHydrogen supply",
     "按工艺化学计量比（约 5:1 过量）向反应器持续供 H₂",
     "PID-XXX Zone B",
     "BT03, FIC-003, PCV, PT"),
    ("N3", "预热器与反应器组 ｜\nPreheater & reactor train",
     "在 60 barG 压力下将进料预热到反应温度并连续氢化",
     "PID-XXX Zone C",
     "E-01, CR01, CR02, CR03, TIC, LIC"),
    ("N4", "冷凝器 ｜\nCondenser",
     "冷却反应器出口物流，凝结液相组分",
     "PID-XXX Zone D",
     "E-02, TIC"),
    ("N5", "气液分离器 ｜\nGas-liquid separator (Demister)",
     "分离气相 H₂ 与液相反应产物；侧入 + 丝网除沫 + 防涡管",
     "PID-XXX Zone E",
     "SE01, LT-101 (guided wave), LSHH, LSLL, CV03, CV08"),
    ("N6", "产品收集 ｜\nProduct collection",
     "收集分离器液相产品，转移至下游",
     "PID-XXX Zone F",
     "BT04, MP03 (transfer pump)"),
    ("N7", "放空与泄放系统 ｜\nVent & relief system",
     "过压保护、正常放空、火焰阻火与紧急泄放",
     "PID-XXX Zone G",
     "PSV-19, PSV-22, PSV-30, RD, flame arrester"),
    ("N8", "公用工程 ｜\nUtilities",
     "导热油循环 (HTF)、氮气钝化 (N₂)、仪表空气 (IA)、电力",
     "PID-XXX Zone H",
     "HTF supply & return, N₂ purge, IA, UPS"),
    ("N9", "取样系统 ｜\nSampling system",
     "在线取样以支持产品分析（当前为 60→0 barG 减压）",
     "PID-XXX Zone I",
     "Sample point SP01, sample bottle system"),
    ("N10", "控制系统 ｜\nControl system (PLC + HMI)",
     "过程控制、报警、联锁、模式切换（正常 / 清洗 / 停车）",
     "Control Philosophy",
     "PLC, HMI, hardwired interlocks, safety instrumented functions"),
    ("N11", "清洗系统 ｜\nCleaning system (CIP)",
     "在批次之间对分离器与反应器进行清洗；含喷淋球与排放",
     "PID-XXX Zone J",
     "Spray balls, drain valves, high-level interlock defeat mechanism"),
]

for i, node in enumerate(nodes, start=2):
    data_row(ws2, i, node, height=48)


# ==================================================================
# Sheet 3: HAZOP Worksheet 工作表 (main)
# ==================================================================
ws3 = wb.create_sheet("3. HAZOP 工作表 Worksheet")

# Column layout
hazop_cols = [
    ("#",                   5),
    ("节点\nNode",           8),
    ("参数\nParameter",       10),
    ("引导词\nGuide word",   10),
    ("偏离\nDeviation",      18),
    ("原因\nCause",          25),
    ("后果\nConsequence",    28),
    ("现有防护\nExisting safeguards", 25),
    ("S",                    5),
    ("L",                    5),
    ("R=S×L",                7),
    ("建议措施\nRecommendation", 30),
    ("负责人\nOwner",         10),
    ("截止\nDue",             10),
    ("状态\nStatus",          10),
]
set_col_widths(ws3, [w for _, w in hazop_cols])
header_row(ws3, 1, [h for h, _ in hazop_cols])

# Pre-populated deviation templates per node
deviations = [
    # (Node, Parameter, Guide word, Deviation description)
    # N1 — Feed & pumps
    ("N1", "流量 Flow",       "无 No",         "计量泵停泵，无原料进入反应器"),
    ("N1", "流量 Flow",       "更少 Less",     "计量泵流量低于设定"),
    ("N1", "流量 Flow",       "更多 More",     "计量泵超流量"),
    ("N1", "流量 Flow",       "反向 Reverse",  "反应器物料回流入原料罐"),
    ("N1", "液位 Level",      "更少 Less",     "原料罐液位过低，泵空转"),
    ("N1", "液位 Level",      "更多 More",     "原料罐过高，可能溢出"),
    ("N1", "组成 Composition","其它 Other than", "误装入错误溶剂或组分"),
    ("N1", "压力 Pressure",   "更多 More",     "泵出口压力过高"),
    # N2 — H2 supply
    ("N2", "流量 Flow",       "无 No",         "H₂ 供给中断"),
    ("N2", "流量 Flow",       "更多 More",     "H₂ 过量进料"),
    ("N2", "流量 Flow",       "更少 Less",     "H₂ 供给不足"),
    ("N2", "压力 Pressure",   "更多 More",     "H₂ 供给超压"),
    ("N2", "压力 Pressure",   "更少 Less",     "H₂ 供给欠压"),
    ("N2", "组成 Composition","其它 Other than","H₂ 纯度不达标或含杂质"),
    ("N2", "泄漏 Leak",       "伴随 As well as", "H₂ 沿管路或密封处泄漏"),
    # N3 — Reactor
    ("N3", "温度 Temperature","更多 More",     "反应器超温 (>150 °C 甚至 300 °C)"),
    ("N3", "温度 Temperature","更少 Less",     "反应器温度过低，反应停止"),
    ("N3", "压力 Pressure",   "更多 More",     "反应器超压 (逼近 80–88 barG PSV set)"),
    ("N3", "液位 Level",      "更多 More",     "反应器液位高，气液比失衡"),
    ("N3", "液位 Level",      "更少 Less",     "反应器空液，催化剂床干燥"),
    ("N3", "反应 Reaction",   "更多 More",     "反应失控 (runaway reaction)"),
    ("N3", "冷却 Cooling",    "无 No",         "冷媒 (HTF) 失效"),
    ("N3", "流量 Flow",       "无 No",         "无进料 + H₂ 继续进入，气相过压"),
    ("N3", "堵塞 Blockage",   "其它 Other than","催化剂床堵塞或管路结垢"),
    # N4 — Condenser
    ("N4", "冷却 Cooling",    "更少 Less",     "冷凝器冷却介质流量不足"),
    ("N4", "冷却 Cooling",    "无 No",         "冷凝器冷却完全失效"),
    ("N4", "压力 Pressure",   "更多 More",     "冷凝器高压侧超压"),
    ("N4", "泄漏 Leak",       "伴随 As well as","换热管破裂 (tube rupture)"),
    # N5 — Separator (critical node)
    ("N5", "液位 Level",      "更多 More",     "SE01 液位过高触发 LSHH"),
    ("N5", "液位 Level",      "更少 Less",     "SE01 液位过低触发 LSLL"),
    ("N5", "压力 Pressure",   "更多 More",     "SE01 气相压力过高"),
    ("N5", "流量 Flow",       "无 No",         "液相无排出 (CV04 卡关)"),
    ("N5", "流量 Flow",       "无 No",         "气相无排出 (CV08 卡关)"),
    ("N5", "仪表 Instrument", "失效 Failure",  "Guided wave 液位变送器失效"),
    ("N5", "相态 Phase",      "其它 Other than","超临界状态出现，泄放公式不适用"),
    ("N5", "清洗 Cleaning",   "其它 Other than","浸泡清洗时高液位联锁被绕过"),
    # N6 — Product collection
    ("N6", "液位 Level",      "更多 More",     "BT04 液位过高 (常压罐可能溢流)"),
    ("N6", "液位 Level",      "更少 Less",     "BT04 空罐，下游转移泵空转"),
    ("N6", "压力 Pressure",   "更多 More",     "常压罐正压 (氮封超压)"),
    ("N6", "压力 Pressure",   "更少 Less",     "常压罐真空 (夜间降温或误抽)"),
    # N7 — Vent / Relief
    ("N7", "PSV 动作 PSV action", "无 No",     "PSV 拒动 (卡涩或结垢)"),
    ("N7", "PSV 动作 PSV action", "更少 Less", "PSV 泄放能力不足 (超临界工况)"),
    ("N7", "排放路径 Vent path",  "堵塞 Blocked","阻火器堵塞导致背压升高"),
    ("N7", "背压 Backpressure",   "更多 More", "vent header 背压 >10% set pressure"),
    ("N7", "反向流动 Reverse flow","反向 Reverse","vent 反流入 BT01 引起 H₂-N₂ 交叉污染"),
    # N8 — Utilities
    ("N8", "HTF",             "无 No",         "导热油系统失效"),
    ("N8", "N₂",              "无 No",         "氮气钝化失效"),
    ("N8", "仪表空气 IA",     "无 No",         "仪表空气失效 → 所有 FC 阀回位"),
    ("N8", "电力 Power",      "无 No",         "断电 (含 UPS 后备失效)"),
    # N9 — Sampling
    ("N9", "取样 Sampling",   "其它 Other than","60→0 barG 减压导致样品闪蒸，样品不代表"),
    ("N9", "取样 Sampling",   "泄漏 Leak",     "取样时管路密封失效，物料泄漏"),
    ("N9", "取样 Sampling",   "错误 Error",    "误操作取样阀导致排放"),
    # N10 — Control
    ("N10", "PLC",            "失效 Failure",  "PLC 主机故障"),
    ("N10", "通讯 Comms",     "失效 Failure",  "PLC-HMI 通讯中断"),
    ("N10", "传感器 Sensor",  "失效 Failure",  "关键传感器故障 (PT, TT, LT)"),
    ("N10", "执行器 Actuator","失效 Failure",  "关键控制阀执行器故障"),
    ("N10", "模式 Mode",      "其它 Other than","操作模式误切换 (清洗→运行时联锁未复位)"),
    # N11 — Cleaning
    ("N11", "清洗 Cleaning",  "更少 Less",     "清洗液流量不足，未达到有效清洗"),
    ("N11", "排放 Drain",     "更少 Less",     "排放速率过慢，无法排空"),
    ("N11", "联锁 Interlock", "失效 Defeat",   "浸泡时高液位联锁被绕过后未复位"),
    ("N11", "液位 Level",     "更多 More",     "液位超过 guided wave 外壳，进入清洁盲区"),
    ("N11", "清洁度 Cleanliness","更少 Less",  "清洗后残余物超验收阈值"),
]

# Fill deviation rows
for i, (node, param, gword, deviation) in enumerate(deviations, start=2):
    r = i
    ws3.cell(row=r, column=1, value=r-1)
    ws3.cell(row=r, column=2, value=node)
    ws3.cell(row=r, column=3, value=param)
    ws3.cell(row=r, column=4, value=gword)
    ws3.cell(row=r, column=5, value=deviation)
    # cause / consequence / safeguards / risk / rec — left blank for team to fill
    for col in range(1, len(hazop_cols) + 1):
        cell = ws3.cell(row=r, column=col)
        cell.font = BODY_FONT
        cell.alignment = LEFT_TOP
        cell.border = BORDER
    # Node # centred
    for col_idx in [1, 2, 3, 4, 9, 10, 11]:
        ws3.cell(row=r, column=col_idx).alignment = CENTER
    ws3.row_dimensions[r].height = 42

# Freeze headers
ws3.freeze_panes = "F2"

# Add data validation for S, L columns (via manual note only — openpyxl basic)
# For clarity add a legend row at very top:
ws3.merge_cells(start_row=len(deviations)+3, start_column=1, end_row=len(deviations)+3, end_column=15)
c = ws3.cell(row=len(deviations)+3, column=1,
             value="S = Severity (1-5) · L = Likelihood (1-5) · R = S × L · "
                   "Status: OPEN / IN PROG / CLOSED · 参考 Sheet 5 风险矩阵")
c.font = Font(name="Microsoft YaHei", size=9, italic=True, color="64748B")
c.alignment = LEFT


# ==================================================================
# Sheet 4: Guide Word & Parameter Reference
# ==================================================================
ws4 = wb.create_sheet("4. 引导词参考 Reference")
set_col_widths(ws4, [14, 14, 55])

# Table 1: Guide words
header_row(ws4, 1, ["引导词\nGuide Word", "英文\nEnglish", "含义\nMeaning"])
guide_words = [
    ("无", "No / None", "设计意图完全没实现（无流量、无温度控制等）"),
    ("更多", "More", "定量参数比设计意图更多（更高流量、更高温度、更高压力）"),
    ("更少", "Less", "定量参数比设计意图更少（较低流量、较低温度）"),
    ("伴随", "As well as", "设计意图完成的同时伴随其它情形（同时发生的次生事件）"),
    ("部分", "Part of", "设计意图只部分完成（一部分组分缺失、部分回路失效）"),
    ("反向", "Reverse", "设计意图相反方向发生（反流、反向反应）"),
    ("其它", "Other than", "非设计意图的情况（错料、错工况、错序）"),
    ("早/晚", "Early / Late", "时序偏差（阀太早/太晚动作）"),
    ("先/后", "Before / After", "顺序偏差（步骤颠倒）"),
]
for i, (cn, en, meaning) in enumerate(guide_words, start=2):
    data_row(ws4, i, [cn, en, meaning])

# Table 2: Parameters
start_row_p = len(guide_words) + 4
ws4.cell(row=start_row_p, column=1, value="参数 Parameters").font = NAVY_BOLD
ws4.merge_cells(start_row=start_row_p, start_column=1, end_row=start_row_p, end_column=3)
ws4.cell(row=start_row_p, column=1).fill = NAVY_LIGHT

header_row(ws4, start_row_p + 1,
           ["参数\nParameter", "英文\nEnglish", "常见应用\nTypical usage"])
parameters = [
    ("流量", "Flow",        "反应物、H₂、公用工程流量"),
    ("压力", "Pressure",     "反应器、分离器、管路压力"),
    ("温度", "Temperature",  "反应温度、冷却介质温度"),
    ("液位", "Level",        "储罐液位、分离器液位"),
    ("组成", "Composition",  "反应物纯度、催化剂状态、杂质"),
    ("相态", "Phase",        "气/液/超临界，含闪蒸"),
    ("反应", "Reaction",     "反应速率、失控、副反应"),
    ("混合", "Mixing",       "液液混合、气液混合"),
    ("腐蚀", "Corrosion",    "材料兼容性、氢脆"),
    ("公用工程", "Utility",   "HTF、N₂、IA、电力"),
    ("清洗", "Cleaning",     "喷淋、浸泡、排放、干燥"),
    ("取样", "Sampling",     "样品收集、代表性、闪蒸"),
    ("控制", "Control",      "PLC 逻辑、联锁、模式"),
    ("时序", "Timing",       "启停顺序、时间偏差"),
]
for i, (cn, en, use) in enumerate(parameters, start=start_row_p + 2):
    data_row(ws4, i, [cn, en, use])


# ==================================================================
# Sheet 5: Risk Matrix 风险矩阵
# ==================================================================
ws5 = wb.create_sheet("5. 风险矩阵 Risk Matrix")
set_col_widths(ws5, [15] + [10]*5 + [30])

# Severity legend
ws5.cell(row=1, column=1, value="严重度 S ｜ Severity (1-5)").font = NAVY_BOLD
ws5.merge_cells("A1:G1")
ws5.cell(row=1, column=1).fill = NAVY_LIGHT

sev_rows = [
    ("S=1", "轻微 Minor",      "无人身伤害；轻微设备损坏；无环境影响"),
    ("S=2", "较小 Low",         "轻伤（可恢复）；设备可修复"),
    ("S=3", "中等 Moderate",   "多人轻伤或 1 人重伤；主要设备损坏；局部环境影响"),
    ("S=4", "严重 High",        "多人重伤或 1 人死亡；重大设备损失；工厂停产"),
    ("S=5", "灾难 Catastrophic", "多人死亡；工厂重大破坏；重大环境事故"),
]
for i, row in enumerate(sev_rows, start=2):
    ws5.cell(row=i, column=1, value=row[0]).font = BODY_BOLD
    ws5.cell(row=i, column=1).alignment = CENTER
    ws5.cell(row=i, column=2, value=row[1]).font = BODY_BOLD
    ws5.merge_cells(start_row=i, start_column=3, end_row=i, end_column=7)
    ws5.cell(row=i, column=3, value=row[2]).font = BODY_FONT
    for col in range(1, 8):
        ws5.cell(row=i, column=col).border = BORDER
        ws5.cell(row=i, column=col).alignment = LEFT

# Likelihood legend
start = 8
ws5.cell(row=start, column=1, value="可能性 L ｜ Likelihood (1-5)").font = NAVY_BOLD
ws5.merge_cells(start_row=start, start_column=1, end_row=start, end_column=7)
ws5.cell(row=start, column=1).fill = NAVY_LIGHT

lik_rows = [
    ("L=1", "极不可能 Very unlikely", "10 年以上 1 次"),
    ("L=2", "不太可能 Unlikely",       "1–10 年 1 次"),
    ("L=3", "可能 Possible",           "1 年 1 次"),
    ("L=4", "较可能 Likely",           "1 年内多次"),
    ("L=5", "几乎肯定 Almost certain","持续或经常"),
]
for i, row in enumerate(lik_rows, start=start + 1):
    ws5.cell(row=i, column=1, value=row[0]).font = BODY_BOLD
    ws5.cell(row=i, column=1).alignment = CENTER
    ws5.cell(row=i, column=2, value=row[1]).font = BODY_BOLD
    ws5.merge_cells(start_row=i, start_column=3, end_row=i, end_column=7)
    ws5.cell(row=i, column=3, value=row[2]).font = BODY_FONT
    for col in range(1, 8):
        ws5.cell(row=i, column=col).border = BORDER
        ws5.cell(row=i, column=col).alignment = LEFT

# Risk = S × L matrix
start2 = 15
ws5.cell(row=start2, column=1, value="风险矩阵 R = S × L ｜ Risk matrix").font = NAVY_BOLD
ws5.merge_cells(start_row=start2, start_column=1, end_row=start2, end_column=7)
ws5.cell(row=start2, column=1).fill = NAVY_LIGHT

# Header: L=1..5
ws5.cell(row=start2 + 1, column=1, value="S ↓ / L →").font = WHITE_BOLD
ws5.cell(row=start2 + 1, column=1).fill = NAVY_FILL
ws5.cell(row=start2 + 1, column=1).alignment = CENTER
ws5.cell(row=start2 + 1, column=1).border = BORDER
for L in range(1, 6):
    c = ws5.cell(row=start2 + 1, column=L + 1, value=f"L={L}")
    c.fill = NAVY_FILL
    c.font = WHITE_BOLD
    c.alignment = CENTER
    c.border = BORDER

# Body: 5 rows × 5 cols
for S in range(1, 6):
    r = start2 + 1 + S
    c = ws5.cell(row=r, column=1, value=f"S={S}")
    c.fill = NAVY_FILL
    c.font = WHITE_BOLD
    c.alignment = CENTER
    c.border = BORDER
    for L in range(1, 6):
        R = S * L
        cell = ws5.cell(row=r, column=L + 1, value=R)
        cell.alignment = CENTER
        cell.border = BORDER
        # Color by risk level
        if R >= 15:
            cell.fill = RED_FILL
            cell.font = WHITE_BOLD
        elif R >= 8:
            cell.fill = ORANGE_LIGHT
            cell.font = BODY_BOLD
        elif R >= 4:
            cell.fill = YELLOW_FILL
            cell.font = BODY_BOLD
        else:
            cell.fill = GREEN_LIGHT
            cell.font = BODY_FONT

# Legend below
r_leg = start2 + 8
legend_items = [
    (RED_FILL,    "R ≥ 15",  "不可接受 · 必须立即整改 (Intolerable)"),
    (ORANGE_LIGHT,"R = 8–14","高风险 · 尽快整改 (High)"),
    (YELLOW_FILL, "R = 4–7", "中风险 · ALARP 层区评估 (Medium)"),
    (GREEN_LIGHT, "R ≤ 3",   "低风险 · 可接受 (Low)"),
]
for i, (fill, label, meaning) in enumerate(legend_items):
    row = r_leg + i
    c = ws5.cell(row=row, column=1, value=label)
    c.fill = fill
    c.font = BODY_BOLD
    c.alignment = CENTER
    c.border = BORDER
    ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    c2 = ws5.cell(row=row, column=2, value=meaning)
    c2.font = BODY_FONT
    c2.alignment = LEFT
    for col in range(1, 8):
        ws5.cell(row=row, column=col).border = BORDER


# ==================================================================
# Sheet 6: Recommendations Log 建议措施
# ==================================================================
ws6 = wb.create_sheet("6. 建议措施 Rec Log")
rec_cols = [
    ("Rec #",             8),
    ("对应 HAZOP 行\nHAZOP row #", 10),
    ("节点\nNode",         8),
    ("建议措施\nRecommendation", 40),
    ("类型\nType",         15),
    ("负责人\nOwner",      12),
    ("截止\nDue",           12),
    ("状态\nStatus",       12),
    ("闭环证据\nClose-out evidence", 25),
]
set_col_widths(ws6, [w for _, w in rec_cols])
header_row(ws6, 1, [h for h, _ in rec_cols])

# Provide 30 empty rows for team to fill
for r in range(2, 32):
    for col in range(1, len(rec_cols) + 1):
        c = ws6.cell(row=r, column=col)
        c.font = BODY_FONT
        c.alignment = LEFT_TOP
        c.border = BORDER
    ws6.row_dimensions[r].height = 26

# Bottom legend for Type
ws6.merge_cells(start_row=33, start_column=1, end_row=33, end_column=9)
c = ws6.cell(row=33, column=1,
             value="Type 分类：DESIGN 设计变更 / PROCEDURE 操作程序 / "
                   "TRAINING 培训 / DOCUMENTATION 文档 / STUDY 补充分析 / OTHER 其它")
c.font = Font(name="Microsoft YaHei", size=9, italic=True, color="64748B")
c.alignment = LEFT

# ==================================================================
# Save
# ==================================================================
import os
out_path = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "internal_HAZOP_template.xlsx",
)
wb.save(out_path)
print(f"Saved: {out_path}")
