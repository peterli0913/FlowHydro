"""
Fill in the 完成情况 (accomplishment) column and suggested 实际得分
(score) column of the mid-year performance evaluation form for 李涛,
based on documented project activities.
"""

import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

src = "李涛绩效.xlsx"
dst = "李涛绩效_已填写.xlsx"
shutil.copy(src, dst)

wb = load_workbook(dst)
ws = wb["李涛自评"]

# Body font that renders both EN and CN
body_font = Font(name="Microsoft YaHei", size=10)
align = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(style="thin", color="BFC4CC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def write(row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = body_font
    c.alignment = align
    c.border = border

# =========================================================================
# KPI 1: Sandwich 项目管理支持 (weight 0.5)
# =========================================================================
# Row 4 — 连续氢化建设项目管理支持  (main focus)
write(4, 15,
    "截至 2026 年上半年，作为国内项目管理支持，协调 CFCT、IEPE 与 UK Sandwich 三方，"
    "系统推动项目整体节奏，主要进展如下：\n\n"
    "• 已推动关闭三方之间 30 余项关键设计议题（含 PFD / P&ID / URS 三大设计文件对齐、"
    "SE01 气液分离器方案闭环、CE-MD 认证范围确定、防爆区域划分方向确定等），主要设计文件已推进至 v0.4-0.5 版本；\n\n"
    "• 建立并稳定运行三方双周技术协作机制，通过双语《会议纪要跟踪表》系统跟踪超过 100 项议题，"
    "对英方问询、技术分歧、设计变更做到全过程可追溯；\n\n"
    "• 有效应对超临界流体安全阀泄放计算、爆炸危险区域划分等前沿技术议题，"
    "组织国内技术团队与 UK 同行建立对等技术对话，避免项目在关键节点上被动等待；\n\n"
    "• 完成 HAZOP 前设计基线对齐工作 —— 汇总国内 33 项待澄清问题与 UK 一方 59 项 Risk Register 合并至统一跟踪，"
    "为 2026 年 8 月中旬正式 HAZOP 铺路；\n\n"
    "• 完成集团 7 月研发述职会议 Sandwich 章节的项目汇报支持（凯总）。\n\n"
    "下半年计划：继续推动 HAZOP 完成、主设备采购启动、CE 认证正式启动，为 2027 年 FAT 与现场开车目标铺路。项目整体节奏可控。"
)
write(4, 16, "0.5")
write(4, 17, "项目按计划推进，超期风险已受控")

# Row 5 — HPLC 和冻干机建设项目管理支持
write(5, 15,
    "已协助推动可行性研究报告启动会顺利召开，配合工程设备部与 SW 方开展初期需求澄清、方案对齐工作。"
    "下半年将持续跟进详细方案讨论与项目实施节奏。"
)
write(5, 16, "0.2")
write(5, 17, "项目处于早期阶段，按计划推进")

# Row 6 — 902 贴建项目管理支持
write(6, 15,
    "已配合完成可行性研究报告启动会阶段工作，参与工程设备部与 SW 方之间的方案讨论与需求对接。"
    "下半年将持续推动项目方案细化与实施准备。"
)
write(6, 16, "0.2")
write(6, 17, "项目处于早期阶段，按计划推进")

# =========================================================================
# KPI 2: Sandwich 相关支持任务 (weight 0.2)
# =========================================================================
write(7, 15,
    "负责 Sandwich 站点的月度概算申请、执行监控与月度回顾工作，"
    "累计完成月概算与月回顾全部按时提交。同时协调组织机器学习研讨小组，"
    "推动集团内技术交流与相关工具试用的落地。"
)
write(7, 16, "0.2")
write(7, 17, "月概算与月回顾按时 100% 完成")

# =========================================================================
# KPI 3: 重点工作 (weight 0.3)
# =========================================================================
write(8, 15,
    "配合集团电算化等生产管理重点工作的推进，"
    "支持相关流程梳理与推行落地，全力配合上级安排的各项专项任务顺利开展。"
)
write(8, 16, "0.3")
write(8, 17, "按上级安排配合推进")

# Adjust row heights so the content is fully visible
ws.row_dimensions[4].height = 250
ws.row_dimensions[5].height = 80
ws.row_dimensions[6].height = 80
ws.row_dimensions[7].height = 80
ws.row_dimensions[8].height = 60

# Widen the 完成情况 (C15) and 备注 (C17) columns for readability
from openpyxl.utils import get_column_letter
ws.column_dimensions[get_column_letter(15)].width = 55
ws.column_dimensions[get_column_letter(16)].width = 10
ws.column_dimensions[get_column_letter(17)].width = 26

wb.save(dst)
print(f"Saved: {dst}")
