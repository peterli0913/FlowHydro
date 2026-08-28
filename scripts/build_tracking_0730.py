"""
Build "Meeting Minutes Tracking Record 0730.xlsx" from the 0729 workbook.

- Adds a new leading sheet "0730" (full carry-forward of "0729")
- Appends the action items arising from the 2026-07-29 biweekly leadership /
  SW review meeting (transcript: "2026-07-2915.52录音.mp3.docx")
- Updates the progress column of the related existing rows with 07/29 notes

New rows are highlighted in column C with amber fill so this batch is easy to
spot (the 07/23 batch used blue FF04B0F1).
"""

import shutil
from copy import copy

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SRC = "Meeting Minutes Tracking Record 0729.xlsx"
DST = "Meeting Minutes Tracking Record 0730.xlsx"

NEW_FILL = "FFFFC000"          # amber — marks the 07/29 meeting batch
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --------------------------------------------------------------------------
# New action items from the 07/29 biweekly leadership + SW review
# (date, originator, theme_cn, theme_en, progress_cn, progress_en, owner)
# --------------------------------------------------------------------------
NEW_ITEMS = [
    (
        "07/29", "SW",
        "针对 R24-CR01 直径变更（DN80→DN65/DN50），建立变更影响日志/清单，逐项列出受影响的"
        "设备、仪表、管线、安全阀及相关文件与参数，确保无遗漏。",
        "Establish a change log / impact list for the R24-CR01 diameter change "
        "(DN80→DN65/DN50), itemising every affected item of equipment, "
        "instrumentation, piping, relief device and the related documents and "
        "parameters, so that nothing is missed.",
        "07/29：会上同意建立变更影响日志，由国内团队逐项梳理受影响范围并跟踪关闭。",
        "07/29: Agreed at the meeting to create a change-impact log; the China "
        "team will work through the affected scope item by item and track it to "
        "closure.",
        "Shaofeng Gao/Chaoqun Hu",
    ),
    (
        "07/29", "SW",
        "反应器 DN65 与 DN50 两种规格并行准备设计方案：先确定处理量，再分别出 PFD 与物料平衡；"
        "并核实仪表选型、安全阀/泄放组合是否需相应调整。待 9 月研发提供 DN65 实验数据后定案。",
        "Prepare two parallel design cases for the DN65 and DN50 reactor: fix "
        "the throughput first, then issue a PFD and mass balance for each, and "
        "verify whether instrument selection and the relief-valve combination "
        "need to change accordingly. To be finalised once R&D provides the DN65 "
        "test data in September.",
        "07/29：会上确定按 DN65 / DN50 两套方案并行准备，使 9 月数据到位后可直接定案，不影响整体进度。",
        "07/29: Agreed to progress both the DN65 and DN50 cases in parallel so "
        "that the decision can be taken as soon as the September data is "
        "available, without affecting the overall schedule.",
        "Shaofeng Gao/Chaoqun Hu/Xiantao Feng",
    ),
    (
        "07/29", "IEPE",
        "容器设计压力由 80 barg 调整为 75 barg，以保持法兰等级 CLASS 900、避免材料选型受限；"
        "需在设备数据表、图纸及相关计算书中同步更新落实。",
        "Vessel design pressure revised from 80 barg to 75 barg in order to "
        "retain the CLASS 900 flange rating and avoid limitations on material "
        "selection; to be updated consistently in the equipment datasheets, "
        "drawings and supporting calculations.",
        "07/29：周二与 Keith 讨论已取得初步一致，继续推进并在各设计文件中落实。",
        "07/29: Initial agreement reached with Keith on Tuesday; to be carried "
        "forward and reflected in the design documents.",
        "Chaoqun Hu/Ziliang Zhao",
    ),
    (
        "07/29", "SW",
        "四撬块分体运输 vs 屋顶整体吊装的决策路径：国内先完成 3D 布置定稿，再以定稿版与 CE 认证机构"
        "评估现场组装后的验证方式、费用与工期；Keith 下周四返回后确认四撬块方案可行性；"
        "SW 侧评估屋顶更换的费用、工期及 R23 停用影响。",
        "Decision path for four-module split shipment vs a whole-unit lift "
        "through the roof: the China team to finalise the 3D layout first, then "
        "use the finalised version to evaluate with the CE Notified Body the "
        "post-assembly verification route, cost and duration; confirm the "
        "four-module option with Keith after he returns next Thursday; SW to "
        "assess the roof-replacement cost, duration and the impact of taking "
        "R23 out of service.",
        "07/29：会上澄清四撬块方案技术上可行，屋顶吊装并非必选项；认证机构评估需以 3D 定稿为前提，"
        "国内加快 3D 定稿以尽早取得费用与工期结论。",
        "07/29: Clarified at the meeting that the four-module option is "
        "technically feasible and roof removal is not mandatory. The Notified "
        "Body assessment depends on the finalised 3D layout, so the China team "
        "will accelerate finalisation to obtain the cost and schedule answer "
        "as early as possible.",
        "Ziliang Zhao/Shaofeng Gao/Keith",
    ),
    (
        "07/29", "SW",
        "明确下阶段（建造 / FAT）文件的编制与签发流程：由谁编制、由谁签发；SW 场地有自有的 FAT / 验证"
        "模板，需与国内模板对比并识别差异。CE 认证方参与文件评审但不参与签发。约定后续安排专题讨论。",
        "Define the preparation and sign-off process for the next-stage "
        "(construction / FAT) documentation — who prepares and who signs off. "
        "The SW site has its own FAT / validation template, which needs to be "
        "compared against the China template so that any gaps are identified. "
        "The CE Notified Body will take part in document review but not in "
        "sign-off. A dedicated follow-up discussion to be arranged.",
        "07/29：双方同意尽早对比两侧模板并识别差异；倾向直接采用 UK 模板以便英国侧验证与确认，"
        "国内按该模板提供资料，避免后期返工。",
        "07/29: Both sides agreed to compare the two templates early and "
        "identify any gaps; the preference is to adopt the UK template so that "
        "verification and validation on the UK side is straightforward, with "
        "the China team supplying information to that template to avoid rework "
        "later.",
        "Shaofeng Gao/Ziliang Zhao",
    ),
    (
        "07/29", "IEPE",
        "请 SW / Keith 提供英国现场验证（validation）所需的信息清单与模板，国内据此提供资料并"
        "按该模板执行，以便尽早识别中英双方要求差异。",
        "SW / Keith to provide the list of information required for the UK site "
        "verification / validation process, together with the template, so that "
        "the China team can supply the information and work to that template "
        "and any differences between the UK and China requirements are "
        "identified early.",
        "07/29：会上确定由 SW 提供模板与信息需求清单，国内提供文件 / 信息并按模板填写；下周与 Keith 沟通落实。",
        "07/29: Agreed that SW will provide the template and the information "
        "requirement list; the China team will supply the documents / "
        "information and complete them to that template. To be taken forward "
        "with Keith next week.",
        "Keith/Shaofeng Gao",
    ),
    (
        "07/29", "SW",
        "HAZOP 前须与 Keith 就 3D 布置与控制说明（含操作手册）达成一致 —— 该两项目前仅国内内部讨论过，"
        "是余下唯二可能引发重大设计变更的环节，须在 HAZOP 前关闭，以确保 HAZOP 只做一次、避免重做。",
        "The 3D layout and the control philosophy (including the operating "
        "manual) must be agreed with Keith before HAZOP. These two items have "
        "so far only been reviewed internally by the China team and are the "
        "only remaining items that could still trigger a significant design "
        "change; they must be closed out beforehand so that HAZOP is run once "
        "only and does not have to be repeated.",
        "07/29：会上确认目标为「HAZOP 只做一次」；两项在 9 月初 HAZOP 前与 Keith 讨论定稿。",
        "07/29: The meeting confirmed the objective of running HAZOP only once; "
        "both items to be discussed and finalised with Keith ahead of the early "
        "September HAZOP.",
        "All",
    ),
]


# --------------------------------------------------------------------------
# Progress notes to append to existing rows (row -> (cn, en))
# --------------------------------------------------------------------------
PROGRESS_UPDATES = {
    120: ("07/29：会上确定按 DN65 / DN50 两套方案并行准备，并建立变更影响日志；"
          "待 9 月研发 DN65 实验数据后最终定案。",
          "07/29: Agreed to prepare the DN65 and DN50 cases in parallel and to "
          "set up a change-impact log; final decision after the September DN65 "
          "R&D test data."),
    70:  ("07/29：会上澄清四撬块方案技术可行、屋顶吊装非必选；认证机构评估以 3D 定稿为前提，"
          "Keith 下周四返回后进一步确认。",
          "07/29: Clarified that the four-module option is technically feasible "
          "and roof removal is not mandatory; the Notified Body assessment "
          "requires the finalised 3D layout, to be confirmed further with Keith "
          "after he returns next Thursday."),
    71:  ("07/29：需以 3D 布置定稿版与 CE 认证机构沟通，评估现场组装后的验证方式、费用与工期。",
          "07/29: The finalised 3D layout is required to engage the CE Notified "
          "Body and assess the post-assembly verification route, cost and "
          "duration."),
    117: ("07/29：操作维护手册预计 8 月中完成。",
          "07/29: The O&M manual is scheduled for completion in mid-August."),
}


def autofit(ws, row):
    """Clear any explicit height so the spreadsheet auto-fits wrapped text."""
    # customHeight is derived from height, so clearing height is sufficient
    ws.row_dimensions[row].height = None


def main():
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    src_ws = wb["0729"]
    new_ws = wb.copy_worksheet(src_ws)
    new_ws.title = "0730"
    # move the new sheet to the front
    wb.move_sheet("0730", offset=-(len(wb.sheetnames) - 1))

    # carry over sheet-level settings that copy_worksheet does not clone
    new_ws.freeze_panes = src_ws.freeze_panes
    for key, dim in src_ws.column_dimensions.items():
        new_ws.column_dimensions[key].width = dim.width
    for idx, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            new_ws.row_dimensions[idx].height = dim.height

    # ---- append progress notes to existing rows -------------------------
    for row, (cn, en) in PROGRESS_UPDATES.items():
        cell = new_ws.cell(row, 4)
        existing = (str(cell.value).rstrip() if cell.value is not None else "")
        cell.value = f"{existing}\n{cn}\n{en}" if existing else f"{cn}\n{en}"
        cell.alignment = Alignment(wrap_text=True, vertical="center",
                                   horizontal="left")
        # the row is now taller than before — let it auto-fit
        autofit(new_ws, row)

    # ---- append the new action items ------------------------------------
    start = new_ws.max_row + 1
    for i, (date, orig, t_cn, t_en, p_cn, p_en, owner) in enumerate(NEW_ITEMS):
        r = start + i
        theme = f"{t_cn}\n{t_en}"
        prog = f"{p_cn}\n{p_en}"
        values = [date, orig, theme, prog, "NO", owner]
        for c, val in enumerate(values, start=1):
            cell = new_ws.cell(r, c, val)
            cell.font = Font(name="Arial", size=10)
            cell.border = BORDER
            if c in (3, 4):
                cell.alignment = Alignment(wrap_text=True, vertical="center",
                                           horizontal="left")
            elif c == 1:
                cell.alignment = Alignment(vertical="center",
                                           horizontal="center")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="center",
                                           horizontal="center")
        # highlight the theme cell so this batch is easy to spot
        new_ws.cell(r, 3).fill = PatternFill("solid", fgColor=NEW_FILL)
        autofit(new_ws, r)

    # keep the autofilter covering the full range
    new_ws.auto_filter.ref = f"A1:G{new_ws.max_row}"

    wb.save(DST)
    print(f"Saved: {DST}")
    print(f"  sheets: {wb.sheetnames}")
    print(f"  0730 rows: {new_ws.max_row} (added {len(NEW_ITEMS)} new items, "
          f"rows {start}-{start + len(NEW_ITEMS) - 1})")
    print(f"  progress updated on rows: {sorted(PROGRESS_UPDATES)}")


if __name__ == "__main__":
    main()
